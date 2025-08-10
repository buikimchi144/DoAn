import sys
import numpy as np
import hashlib
import openpyxl
from datetime import datetime
import pyodbc
import os
from db.database import Database
from ui_components import CustomTableWidget, Sidebar
from face_recognition_util import FaceRecognitionUtil
import time
from camera import WebcamThread
from camera import Camera
import base64
import cv2
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import QWidget, QPushButton, QLabel, QVBoxLayout, QHBoxLayout, QFrame, QGraphicsDropShadowEffect
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
from ui_components import CustomButton, CustomTableWidget
import pandas as pd

class EmployeeAttendanceApp(QMainWindow):
    def __init__(self, user_info=None, db=None, face_recognizer=None, controller_window=None):
        super().__init__()
        self.user_info = user_info or {}
        self.setWindowTitle("Hệ thống điểm danh nhân viên")
        self.setGeometry(100, 50, 1000, 700)
        self.setWindowIcon(QIcon("./assets/img/dashboard.png"))
        self.setStyleSheet("QMainWindow { background: #f5f5f5; }")

        # Initialize components
        self.db = db
        self.current_frame = None
        self.current_embedding = None
        self.face_util = face_recognizer # new update
        self.controller_window = controller_window
        self.timer = None
        self.current_cam = None
        self.edit_mode = None  # 'employee' hoặc 'attendance'
        self.edit_data = None  # Dữ liệu đang được edit

        # Danh sách lưu các frame để tự động chọn avatar tốt nhất
        self.captured_frames = []
        self.selected_avatar_frame = None
        self.face_qualities = []  # Lưu chất lượng khuôn mặt của từng frame

        self.init_ui()

    def init_ui(self):
        # Sidebar với callbacks
        callbacks = {
            "show_employee_list": self.show_employee_list,
            "show_register": self.show_register,
            "show_attendance_stats": self.show_attendance_stats,
            "logout": self.logout,
        }
        sidebar = Sidebar(callbacks)

        # Header đơn giản
        header = QFrame()
        header.setFixedHeight(60)
        header.setStyleSheet("QFrame { background: #2196F3; }")

        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(20, 0, 20, 0)

        self.page_title = QLabel("DANH SÁCH NHÂN VIÊN")
        self.page_title.setFont(QFont("Arial", 16, QFont.Bold))
        self.page_title.setStyleSheet("color: white;")

        user_info = QLabel("Admin")
        user_info.setStyleSheet("color: white; padding: 5px 10px; border: 1px solid white; border-radius: 15px;")

        header_layout.addWidget(self.page_title)
        header_layout.addStretch()
        header_layout.addWidget(user_info)

        # Main content
        self.main_frame = QFrame()
        self.main_frame.setStyleSheet("QFrame { background: white; border-radius: 8px; }")

        self.main_layout = QVBoxLayout()
        self.main_layout.setContentsMargins(20, 20, 20, 20)
        self.main_frame.setLayout(self.main_layout)

        # Layout chính
        content_layout = QVBoxLayout()
        content_layout.setSpacing(10)
        content_layout.setContentsMargins(10, 10, 10, 10)
        content_layout.addWidget(header)
        content_layout.addWidget(self.main_frame)

        content_widget = QWidget()
        content_widget.setLayout(content_layout)

        main_widget = QWidget()
        main_layout = QHBoxLayout()
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(sidebar)
        main_layout.addWidget(content_widget, 1)
        main_widget.setLayout(main_layout)

        self.setCentralWidget(main_widget)
        self.show_register()

    def logout(self):
        # Clean up camera if running
        self.cleanup_camera()
        self.hide()
        if self.controller_window:
            self.controller_window.show_main_window()  # Gọi phương thức đã sửa
        else:
            print("❌ controller_window is None!")

    def cleanup_camera(self):
        """Clean up camera resources"""
        if self.timer and self.timer.isActive():
            self.timer.stop()
        if self.current_cam:
            self.current_cam.release()
            self.current_cam = None

    def show_employee_list(self):
        self.cleanup_camera()  # Clean up any running camera
        self.clear_layout(self.main_layout)
        self.page_title.setText("DANH SÁCH NHÂN VIÊN")
        self.build_employee_list_ui()

    def build_employee_list_ui(self):
        """
        Xây dựng giao diện cho danh sách nhân viên, sử dụng các component đã tối ưu.
        """
        self.clear_layout(self.main_layout)  # Giả định bạn có hàm này để xóa widget cũ

        try:
            # Lấy dữ liệu từ database
            self.all_employees = self.db.employees.get_all_employees()
            self.filtered_employees = self.all_employees.copy()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi truy vấn dữ liệu nhân viên: {str(e)}")
            return

        # 1. Layout cho các thẻ thống kê
        stats_layout = QHBoxLayout()
        # Giả định bạn có hàm create_stat_card trong class chính
        self.total_stat = self.create_stat_card("Tổng NV", str(len(self.all_employees)), "#4CAF50")
        self.active_stat = self.create_stat_card("Hiển thị", str(len(self.filtered_employees)), "#2196F3")
        stats_layout.addWidget(self.total_stat)
        stats_layout.addWidget(self.active_stat)
        stats_layout.addStretch()

        # 2. Layout cho các nút hành động (thay thế cho RefreshButtonComponent)
        action_buttons_layout = QHBoxLayout()

        # Sử dụng trực tiếp CustomButton đã được tối ưu
        refresh_btn = CustomButton("🔄 Làm mới", button_type="primary")
        refresh_btn.clicked.connect(self.build_employee_list_ui)

        action_buttons_layout.addWidget(refresh_btn)
        action_buttons_layout.addStretch()  # Đẩy nút về bên trái

        # 3. Layout cho thanh tìm kiếm
        # Giả định bạn có hàm create_search_layout trong class chính
        search_layout = self.create_search_layout()

        # 4. Bảng dữ liệu nhân viên
        headers = ["ID", "Họ và tên", "Phòng ban", "Giới tính", "Chức vụ", "Ngày sinh", "Ngày vào", "Quyền"]

        # Sử dụng CustomTableWidget đã được tối ưu
        self.employee_table = CustomTableWidget(
            headers=headers,
            data=self.filtered_employees,
            # Không cần nút "Thêm" trong bảng này, nên để add_button_text là None (mặc định)
        )
        self.employee_table.edit_clicked.connect(self.handle_edit_employee)
        self.employee_table.delete_clicked.connect(self.handle_delete_employee)

        # Thêm tất cả các layout và widget vào layout chính
        self.main_layout.addLayout(stats_layout)
        self.main_layout.addLayout(action_buttons_layout)
        self.main_layout.addLayout(search_layout)
        self.main_layout.addWidget(self.employee_table)
    def _create_section_group_box(self, title):
        """Tạo một QGroupBox với tiêu đề và style cho các phần tìm kiếm/lọc"""
        group_box = QGroupBox(title)
        group_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: 600;
                color: #334155;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 10px; /* Space for the title */
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 10px;
                margin-left: 5px;
                color: #1e293b;
            }
        """)
        group_box.setLayout(QVBoxLayout())  # Use QVBoxLayout for the content
        group_box.layout().setContentsMargins(10, 20, 10, 10)  # Adjust margins
        return group_box

    def export_to_excel(self):
        """Xuất dữ liệu nhân viên đang hiển thị ra file Excel."""
        if not self.filtered_employees:
            QMessageBox.information(self, "Thông báo", "Không có dữ liệu để xuất Excel.")
            return

        # Lấy đường dẫn lưu file
        file_path, _ = QFileDialog.getSaveFileName(self, "Lưu file Excel", "danh_sach_nhan_vien.xlsx",
                                                   "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            # Lấy header từ bảng
            headers = ["ID", "Họ và tên", "Phòng ban", "Giới tính", "Chức vụ", "Ngày sinh", "Ngày vào", "Quyền"]

            # --- DEBUGGING PRINTS (Keep for a final check, then you can remove them) ---
            print(f"Type of self.filtered_employees: {type(self.filtered_employees)}")
            print(f"Number of rows in self.filtered_employees: {len(self.filtered_employees)}")
            if isinstance(self.filtered_employees, list) and self.filtered_employees:
                first_element_raw = self.filtered_employees[0]
                print(f"First element in filtered_employees (raw): {first_element_raw}")
                print(f"Type of first element (raw): {type(first_element_raw)}")
                print(f"Length of first element (raw): {len(first_element_raw)}")
            # --- END DEBUGGING PRINTS ---

            # IMPORTANT FIX: Convert each pyodbc.Row object to a standard tuple
            # This list comprehension iterates through each pyodbc.Row object
            # and converts it into a standard Python tuple.
            processed_employees_data = [tuple(row_obj) for row_obj in self.filtered_employees]

            # Tạo DataFrame từ dữ liệu đã lọc
            df = pd.DataFrame(processed_employees_data, columns=headers)  # Use the converted data here

            # Ghi DataFrame ra file Excel
            df.to_excel(file_path, index=False)
            QMessageBox.information(self, "Thành công", f"Dữ liệu đã được xuất ra:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Không thể xuất file Excel: {str(e)}")

    def populate_unit_combo(self):
        """Điền dữ liệu cho combo box Chức vụ"""
        units = set()
        for emp in self.all_employees:
            if len(emp) > 2 and emp[2]:  # Assuming unit is at index 2
                units.add(str(emp[2]))

        self.unit_combo.clear()
        self.unit_combo.addItem("Tất cả")
        self.unit_combo.addItems(sorted(units))

    def populate_role_combo(self):
        """Điền dữ liệu cho combo box quyền"""
        roles = set()
        for emp in self.all_employees:
            if len(emp) > 7 and emp[7]:  # Assuming role is at index 7
                roles.add(str(emp[7]))

        self.role_combo.clear()
        self.role_combo.addItem("Tất cả")
        self.role_combo.addItems(sorted(roles))

    def filter_employees(self):
        """Lọc nhân viên theo các điều kiện tìm kiếm"""
        search_text = self.search_input.text().lower().strip()
        selected_unit = self.unit_combo.currentText()
        selected_gender = self.gender_combo.currentText()
        selected_role = self.role_combo.currentText()
        age_from = self.age_from_spin.value()
        age_to = self.age_to_spin.value()
        year_from = self.year_from_spin.value()
        year_to = self.year_to_spin.value()

        self.filtered_employees = []

        for emp in self.all_employees:
            # Kiểm tra điều kiện tìm kiếm tổng quát
            if search_text:
                ma_nhan_vien = str(emp[0]).lower() if len(emp) > 0 and emp[0] else ""
                ho_ten = str(emp[1]).lower() if len(emp) > 1 and emp[1] else ""

                # Tìm kiếm trong mã nhân viên HOẶC họ tên
                if search_text not in ma_nhan_vien and search_text not in ho_ten:
                    continue

            # Kiểm tra Chức vụ
            if selected_unit != "Tất cả":
                if len(emp) <= 2 or str(emp[2]) != selected_unit:
                    continue

            # Kiểm tra giới tính
            if selected_gender != "Tất cả":
                if len(emp) <= 3 or str(emp[3]) != selected_gender:
                    continue

            # Kiểm tra quyền
            if selected_role != "Tất cả":
                if len(emp) <= 7 or str(emp[7]) != selected_role:
                    continue

            # Kiểm tra độ tuổi (giả sử ngày sinh ở index 5)
            if len(emp) > 5 and emp[5]:
                try:
                    from datetime import datetime
                    birth_date = datetime.strptime(str(emp[5]), "%Y-%m-%d")
                    age = (datetime.now() - birth_date).days // 365
                    if age < age_from or age > age_to:
                        continue
                except:
                    pass  # Bỏ qua nếu không parse được ngày sinh

            # Kiểm tra năm vào làm (giả sử ngày vào ở index 6)
            if len(emp) > 6 and emp[6]:
                try:
                    from datetime import datetime
                    join_date = datetime.strptime(str(emp[6]), "%Y-%m-%d")
                    join_year = join_date.year
                    if join_year < year_from or join_year > year_to:
                        continue
                except:
                    pass  # Bỏ qua nếu không parse được ngày vào

            # Nếu qua hết các điều kiện thì thêm vào kết quả
            self.filtered_employees.append(emp)

        # Cập nhật bảng và thống kê
        self.update_table_and_stats()

        # Sửa lại dòng này - thêm 2 tham số
        self.update_result_info(len(self.filtered_employees), len(self.all_employees))
    def update_table_and_stats(self):
        """Cập nhật bảng và thống kê sau khi lọc"""
        # Cập nhật bảng
        self.employee_table.refresh_data(self.filtered_employees)

        # Cập nhật thống kê
        self.update_stat_card(self.active_stat, "Hiển thị", str(len(self.filtered_employees)), "#2196F3")

    def update_stat_card(self, card_widget, title, value, color):
        """Cập nhật giá trị của stat card"""
        # Nếu bạn có method update_value() trong stat card, sử dụng nó
        if hasattr(card_widget, 'update_value'):
            card_widget.update_value(value)
        else:
            # Fallback: Tìm và cập nhật QLabel chứa số
            for child in card_widget.findChildren(QLabel):
                text = child.text()
                if text.isdigit() or (text.replace(',', '').replace('.', '').isdigit()):
                    child.setText(value)
                    break

    def clear_search(self):
        """Xóa tất cả điều kiện tìm kiếm"""
        self.search_input.clear()
        self.unit_combo.setCurrentIndex(0)
        self.gender_combo.setCurrentIndex(0)
        self.role_combo.setCurrentIndex(0)
        self.age_from_spin.setValue(18)
        self.age_to_spin.setValue(70)
        self.year_from_spin.setValue(2000)
        self.year_to_spin.setValue(2025)
        self.update_result_info(len(self.filtered_employees), len(self.all_employees))

    def create_search_layout(self):
        """Tạo giao diện tìm kiếm với các nút trong khung frame"""

        # Main container với styling đẹp - GIẢM PADDING
        search_group = QGroupBox("🔍 Bộ lọc tìm kiếm")
        search_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 1px solid #d0d7de;
                border-radius: 12px;
                margin-top: 1ex;
                padding-top: 10px;
                background-color: #f8fafc;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 8px 0 8px;
                color: #0969da;
                background-color: #f8fafc;
            }
        """)

        # Layout chính của GroupBox - GIẢM PADDING
        group_main_layout = QVBoxLayout()
        group_main_layout.setContentsMargins(15, 20, 15, 15)
        group_main_layout.setSpacing(10)

        # Layout dòng đầu - CHỈ 1 DÒNG (các bộ lọc)
        filters_layout = QHBoxLayout()
        filters_layout.setSpacing(20)

        # Style chung cho input - Cập nhật font-size nhỏ hơn
        input_style = """
            QLineEdit, QComboBox, QSpinBox, QDateEdit {
                padding: 6px 10px;
                border: 1px solid #d1d5db;
                border-radius: 8px;
                background-color: white;
                font-size: 12px;
                min-height: 14px;
                min-width: 100px;
            }
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDateEdit:focus {
                border-color: #3b82f6;
                box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 4px solid #6b7280;
                margin-right: 4px;
            }
            QLineEdit::placeholder {
                color: #9ca3af;
                font-style: italic;
            }
        """

        # Style cho label - Cập nhật font-size nhỏ hơn
        label_style = """
            QLabel {
                color: #374151;
                font-weight: 500;
                font-size: 12px;
                margin-right: 8px;
            }
        """

        # Helper function để tạo field với label trên CÙNG MỘT DÒNG
        def create_field_group(label_text, widget):
            container = QHBoxLayout()
            container.setSpacing(4)

            label = QLabel(label_text)
            label.setStyleSheet(label_style)
            container.addWidget(label)

            widget.setStyleSheet(input_style)
            container.addWidget(widget)

            return container

        # Tìm kiếm
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Nhập tên, ID")
        self.search_input.textChanged.connect(self.filter_employees)
        search_group_layout = create_field_group("Nhân viên:", self.search_input)
        filters_layout.addLayout(search_group_layout)

        # Phòng ban
        self.unit_combo = QComboBox()
        self.populate_unit_combo()
        self.unit_combo.currentTextChanged.connect(self.filter_employees)
        unit_group_layout = create_field_group("Phòng ban:", self.unit_combo)
        filters_layout.addLayout(unit_group_layout)

        # Giới tính
        self.gender_combo = QComboBox()
        self.gender_combo.addItems(["Tất cả", "Nam", "Nữ"])
        self.gender_combo.currentTextChanged.connect(self.filter_employees)
        gender_group_layout = create_field_group("Giới tính:", self.gender_combo)
        filters_layout.addLayout(gender_group_layout)

        # Vai trò
        self.role_combo = QComboBox()
        self.populate_role_combo()
        self.role_combo.currentTextChanged.connect(self.filter_employees)
        role_group_layout = create_field_group("Vai trò:", self.role_combo)
        filters_layout.addLayout(role_group_layout)

        # Tuổi
        self.age_from_spin = QSpinBox()
        self.age_from_spin.setRange(18, 70)
        self.age_from_spin.setValue(18)
        self.age_from_spin.valueChanged.connect(self.filter_employees)
        age_group_layout = create_field_group("Tuổi:", self.age_from_spin)
        filters_layout.addLayout(age_group_layout)

        # Thêm dòng bộ lọc vào layout chính
        group_main_layout.addLayout(filters_layout)

        # Dòng chứa buttons và thông tin kết quả - TRONG KHUNG FRAME
        button_result_layout = QHBoxLayout()
        button_result_layout.setSpacing(12)

        # Nút Xóa bộ lọc (bên trái) - GIẢM PADDING
        clear_btn = QPushButton("🗑️ Xóa bộ lọc")
        clear_btn.setStyleSheet(self.get_compact_clear_button_style())
        clear_btn.clicked.connect(self.clear_search)
        button_result_layout.addWidget(clear_btn)

        # Nút Xuất Excel - GIẢM PADDING
        export_excel_btn = QPushButton("📊 Xuất Excel")
        export_excel_btn.setStyleSheet(self.get_compact_export_button_style())
        export_excel_btn.clicked.connect(self.export_to_excel)
        button_result_layout.addWidget(export_excel_btn)

        # Khoảng cách đẩy thông tin kết quả sang phải
        button_result_layout.addStretch()

        # Tính toán số liệu thực tế (bên phải)
        if hasattr(self, 'all_employees') and self.all_employees:
            total_count = len(self.all_employees)
            displayed_count = len(getattr(self, 'filtered_employees', self.all_employees))
            result_text = f"Hiển thị {displayed_count}/{total_count} bản ghi"
        else:
            result_text = "Chưa có dữ liệu"

        self.result_info = QLabel(result_text)
        self.result_info.setStyleSheet("""
            QLabel {
                color: #3b82f6;
                font-weight: 600;
                font-size: 12px;
                padding: 6px 10px;
                background-color: #eff6ff;
                border: 1px solid #dbeafe;
                border-radius: 8px;
            }
        """)
        button_result_layout.addWidget(self.result_info)

        # Thêm dòng buttons vào layout chính của GroupBox
        group_main_layout.addLayout(button_result_layout)

        # Thiết lập layout chính cho GroupBox
        search_group.setLayout(group_main_layout)

        # Các trường còn lại ẩn đi nhưng vẫn tạo để không lỗi code
        self.age_to_spin = QSpinBox()
        self.age_to_spin.setRange(18, 70)
        self.age_to_spin.setValue(70)
        self.age_to_spin.valueChanged.connect(self.filter_employees)
        self.age_to_spin.hide()

        self.year_from_spin = QSpinBox()
        self.year_from_spin.setRange(2000, 2025)
        self.year_from_spin.setValue(2000)
        self.year_from_spin.valueChanged.connect(self.filter_employees)
        self.year_from_spin.hide()

        self.year_to_spin = QSpinBox()
        self.year_to_spin.setRange(2000, 2025)
        self.year_to_spin.setValue(2025)
        self.year_to_spin.valueChanged.connect(self.filter_employees)
        self.year_to_spin.hide()

        # Layout tổng - CHỈ RETURN GROUPBOX
        bottom_layout = QVBoxLayout()
        bottom_layout.setSpacing(15)
        bottom_layout.addWidget(search_group)

        return bottom_layout
    def update_result_info(self, shown_count, total_count):
        """Cập nhật thông tin số lượng bản ghi hiển thị"""
        if hasattr(self, 'result_info'):
            self.result_info.setText(f"Hiển thị {shown_count}/{total_count} bản ghi")
    def populate_unit_combo(self):
        """Điền dữ liệu cho combo box Phòng ban"""
        units = set()
        for emp in self.all_employees:
            if len(emp) > 2 and emp[2]:  # Assuming unit is at index 2
                units.add(str(emp[2]))

        self.unit_combo.clear()
        self.unit_combo.addItem("Tất cả")
        self.unit_combo.addItems(sorted(units))

    def update_table_and_stats(self):
        """Cập nhật bảng và thống kê sau khi lọc"""
        # Cập nhật bảng
        self.employee_table.refresh_data(self.filtered_employees)

        # Cập nhật thống kê
        self.update_stat_card(self.active_stat, "Hiển thị", str(len(self.filtered_employees)), "#2196F3")
    def clear_search(self):
        """Xóa tất cả điều kiện tìm kiếm"""
        self.search_input.clear()
        self.unit_combo.setCurrentIndex(0)
        self.gender_combo.setCurrentIndex(0)
        self.role_combo.setCurrentIndex(0)
        self.age_from_spin.setValue(18)
        self.age_to_spin.setValue(70)
        self.year_from_spin.setValue(2000)
        self.year_to_spin.setValue(2025)
    def handle_edit_employee(self, row_index):
        """Xử lý khi click nút Edit employee"""
        try:
            employees = self.db.employees.get_all_employees()
            if row_index < len(employees):
                employee = employees[row_index]

                # Set edit mode
                self.edit_mode = 'employee'
                self.edit_data = {
                    'employee_id': employee[0],  # EmployeeID
                    'full_name': employee[1],  # FullName
                    'department': employee[2],  # Department
                    'gender': employee[3],  # Gender
                    'position': employee[4],  # Position
                    'dob': employee[5],  # DateOfBirth
                    'join_date': employee[6],  # JoinDate
                    'role': employee[7],  # JoinDate
                    'row_index': row_index
                }

                # Chuyển sang form edit
                self.show_edit_employee_form()

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi edit: {str(e)}")
    def show_edit_employee_form(self):
        """Hiển thị form edit employee với camera để cập nhật khuôn mặt"""
        self.cleanup_camera()
        self.clear_layout(self.main_layout)
        self.page_title.setText("SỬA THÔNG TIN NHÂN VIÊN")

        # Main layout với form bên trái, camera bên phải
        main_layout = QHBoxLayout()

        # === FORM LAYOUT (Left side) ===
        form_layout = QGridLayout()
        form_layout.setVerticalSpacing(15)
        form_layout.setHorizontalSpacing(10)

        # Style cho input
        input_style = """
            QLineEdit, QComboBox, QDateEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 12px;
                min-width: 200px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border-color: #2196F3;
            }
        """

        def create_label(text):
            lbl = QLabel(text)
            lbl.setStyleSheet("color: #333; font-weight: bold;")
            return lbl

        # Input fields với dữ liệu hiện tại
        self.edit_emp_name = QLineEdit(self.edit_data['full_name'])
        self.edit_emp_department = QLineEdit(self.edit_data['department'])

        self.edit_emp_gender = QComboBox()
        self.edit_emp_gender.addItems(["Nam", "Nữ", "Khác"])
        gender_index = self.edit_emp_gender.findText(self.edit_data['gender'])
        if gender_index >= 0:
            self.edit_emp_gender.setCurrentIndex(gender_index)

        self.edit_user_role = QComboBox()
        self.edit_user_role.addItems(["Admin", "Nhân viên"])

        # Ánh xạ role từ DB sang hiển thị
        if self.edit_data['role'].lower() == "admin":
            self.edit_user_role.setCurrentIndex(0)
        else:
            self.edit_user_role.setCurrentIndex(1)
        role_index = self.edit_user_role.findText(self.edit_data['role'])
        if role_index >= 0:
            self.edit_user_role.setCurrentIndex(role_index)

        self.edit_emp_position = QLineEdit(self.edit_data['position'])

        self.edit_emp_dob = QDateEdit()
        self.edit_emp_dob.setCalendarPopup(True)
        self.edit_emp_dob.setDisplayFormat("dd/MM/yyyy")
        self.edit_emp_dob.setDate(QDate.fromString(self.edit_data['dob'], "yyyy-MM-dd"))

        self.edit_emp_joindate = QDateEdit()
        self.edit_emp_joindate.setCalendarPopup(True)
        self.edit_emp_joindate.setDisplayFormat("dd/MM/yyyy")
        self.edit_emp_joindate.setDate(QDate.fromString(self.edit_data['join_date'], "yyyy-MM-dd"))

        # Apply styles
        for widget in [self.edit_emp_name, self.edit_emp_department, self.edit_emp_gender,
                       self.edit_user_role, self.edit_emp_position, self.edit_emp_dob, self.edit_emp_joindate]:
            widget.setStyleSheet(input_style)

        form_layout.setSpacing(12)
        # Add to form
        form_layout.addWidget(create_label("Họ và tên:"), 0, 0)
        form_layout.addWidget(self.edit_emp_name, 0, 1)
        form_layout.addWidget(create_label("Phòng ban:"), 1, 0)
        form_layout.addWidget(self.edit_emp_department, 1, 1)
        form_layout.addWidget(create_label("Giới tính:"), 2, 0)
        form_layout.addWidget(self.edit_emp_gender, 2, 1)
        form_layout.addWidget(create_label("Quyền hạn:"), 3, 0)
        form_layout.addWidget(self.edit_user_role, 3, 1)
        form_layout.addWidget(create_label("Chức vụ:"), 4, 0)
        form_layout.addWidget(self.edit_emp_position, 4, 1)
        form_layout.addWidget(create_label("Ngày sinh:"), 5, 0)
        form_layout.addWidget(self.edit_emp_dob, 5, 1)
        form_layout.addWidget(create_label("Ngày vào làm:"), 6, 0)
        form_layout.addWidget(self.edit_emp_joindate, 6, 1)

        # Label chữ "Thông báo:"
        self.edit_label_info = QLabel("Thông báo:")
        self.edit_label_info.setStyleSheet(
            "color: #2196F3; font-weight: bold; padding: 4px 8px; font-size: 12px;"
        )
        self.edit_label_info.setMaximumHeight(30)

        # Label để hiển thị nội dung notice động
        self.edit_label_notice = QLabel("Có thể cập nhật khuôn mặt mới hoặc giữ nguyên")
        self.edit_label_notice.setStyleSheet(
            "color: #2196F3; padding: 4px 8px; background: #e3f2fd; border-radius: 4px; font-size: 11px;"
        )
        self.edit_label_notice.setMaximumHeight(30)
        self.edit_label_notice.setVisible(True)

        form_layout.addWidget(self.edit_label_info, 7, 0)
        form_layout.addWidget(self.edit_label_notice, 7, 1)

        # Buttons
        button_style = """
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background: #1976D2; }
        """

        # Face recognition buttons
        btn_capture_edit = QPushButton("🔄 Cập nhật khuôn mặt")
        btn_capture_edit.setStyleSheet(button_style.replace("#2196F3", "#4CAF50").replace("#1976D2", "#45a049"))
        btn_capture_edit.clicked.connect(self.capture_face_for_edit)

        btn_stop_edit = QPushButton("⏹️ Dừng camera")
        btn_stop_edit.setStyleSheet(button_style.replace("#2196F3", "#FF9800").replace("#1976D2", "#F57C00"))
        btn_stop_edit.clicked.connect(self.stop_camera)

        # Save/Cancel buttons
        btn_save = QPushButton("💾 Lưu thay đổi")
        btn_save.setStyleSheet(button_style)
        btn_save.clicked.connect(self.save_employee_edit)

        btn_cancel = QPushButton("❌ Hủy")
        btn_cancel.setStyleSheet(button_style.replace("#2196F3", "#757575").replace("#1976D2", "#616161"))
        btn_cancel.clicked.connect(self.show_employee_list)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(btn_capture_edit)
        btn_layout.addWidget(btn_stop_edit)
        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_cancel)

        form_layout.addLayout(btn_layout, 8, 0, 1, 2)

        # === CAMERA LAYOUT (Right side) ===
        camera_layout = QVBoxLayout()

        # Current avatar display
        current_avatar_label = QLabel("Ảnh đại diện hiện tại:")
        current_avatar_label.setStyleSheet("color: #333; font-weight: bold; margin-bottom: 5px;")

        self.edit_current_avatar = QLabel("Đang tải...")
        self.edit_current_avatar.setFixedSize(200, 200)
        self.edit_current_avatar.setAlignment(Qt.AlignCenter)
        self.edit_current_avatar.setStyleSheet("border: 2px solid #4CAF50; border-radius: 8px; background: #f5f5f5;")

        # Load current avatar
        self.load_current_avatar()

        # Live camera preview
        camera_title = QLabel("Camera trực tiếp:")
        camera_title.setStyleSheet("color: #333; font-weight: bold; margin-bottom: 5px; margin-top: 10px;")

        self.edit_camera_preview = QLabel("Camera sẽ hiển thị ở đây")
        self.edit_camera_preview.setFixedSize(320, 240)
        self.edit_camera_preview.setAlignment(Qt.AlignCenter)
        self.edit_camera_preview.setStyleSheet(
            "border: 2px solid #2196F3; border-radius: 8px; background: #f5f5f5; color: #666;")

        # New avatar preview
        new_avatar_label = QLabel("Ảnh đại diện mới (nếu có):")
        new_avatar_label.setStyleSheet("color: #333; font-weight: bold; margin-top: 10px;")

        self.edit_new_avatar = QLabel("Chưa chụp ảnh mới")
        self.edit_new_avatar.setFixedSize(200, 200)
        self.edit_new_avatar.setAlignment(Qt.AlignCenter)
        self.edit_new_avatar.setStyleSheet("border: 2px dashed #ccc; border-radius: 8px; color: #666;")

        camera_layout.addWidget(current_avatar_label)
        camera_layout.addWidget(self.edit_current_avatar)
        camera_layout.addWidget(camera_title)
        camera_layout.addWidget(self.edit_camera_preview)
        camera_layout.addWidget(new_avatar_label)
        camera_layout.addWidget(self.edit_new_avatar)
        camera_layout.addStretch()

        # Setup main layout
        form_widget = QWidget()
        form_widget.setLayout(form_layout)

        camera_widget = QWidget()
        camera_widget.setLayout(camera_layout)

        main_layout.addWidget(form_widget, 2)  # Form chiếm 2/3
        main_layout.addWidget(camera_widget, 1)  # Camera chiếm 1/3

        self.main_layout.addLayout(main_layout)

    def create_stat_card(self, title, value, color):
        card = QFrame()
        card.setFixedSize(150, 60)
        card.setStyleSheet(f"QFrame {{ background: white; border-left: 3px solid {color}; border-radius: 5px; }}")

        layout = QHBoxLayout(card)
        layout.setContentsMargins(10, 5, 10, 5)

        text_layout = QVBoxLayout()
        text_layout.setSpacing(0)

        title_label = QLabel(title)
        title_label.setStyleSheet("color: #666; font-size: 10px;")

        value_label = QLabel(value)
        value_label.setFont(QFont("Arial", 14, QFont.Bold))
        value_label.setStyleSheet(f"color: {color};")

        text_layout.addWidget(title_label)
        text_layout.addWidget(value_label)
        layout.addLayout(text_layout)

        return card

    def build_register_form(self):
        """Build registration form with integrated camera - removed manual avatar selection"""
        self.clear_layout(self.main_layout)

        # Form layout đơn giản
        form_layout = QGridLayout()
        form_layout.setVerticalSpacing(15)
        form_layout.setHorizontalSpacing(10)

        # Style đơn giản cho input
        input_style = """
            QLineEdit, QComboBox, QDateEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 12px;
                min-width: 200px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {
                border-color: #2196F3;
            }
        """

        # Input fields
        self.entry_name = QLineEdit()
        self.entry_name.setPlaceholderText("Họ và tên")
        self.department = QLineEdit()
        self.department.setPlaceholderText("Phòng ban")
        self.entry_gender = QComboBox()
        self.entry_gender.addItems(["Nam", "Nữ", "Khác"])
        self.entry_position = QLineEdit()
        self.entry_position.setPlaceholderText("Chức vụ")

        # Thêm role selection
        self.entry_role = QComboBox()
        self.entry_role.addItems(["Nhân viên", "Admin"])

        self.entry_dob = QDateEdit()
        self.entry_dob.setCalendarPopup(True)
        self.entry_dob.setDisplayFormat("dd/MM/yyyy")
        self.entry_joindate = QDateEdit()
        self.entry_joindate.setCalendarPopup(True)
        self.entry_joindate.setDisplayFormat("dd/MM/yyyy")

        for widget in [self.entry_name, self.department, self.entry_gender,
                       self.entry_position, self.entry_role, self.entry_dob, self.entry_joindate]:
            widget.setStyleSheet(input_style)

        # Labels đơn giản
        def create_label(text):
            lbl = QLabel(text)
            lbl.setStyleSheet("color: #333; font-weight: bold;")
            return lbl

        # Add to form
        form_layout.setSpacing(12)
        form_layout.addWidget(create_label("Họ và tên:"), 0, 0)
        form_layout.addWidget(self.entry_name, 0, 1)
        form_layout.addWidget(create_label("Phòng ban:"), 1, 0)
        form_layout.addWidget(self.department, 1, 1)
        form_layout.addWidget(create_label("Giới tính:"), 2, 0)
        form_layout.addWidget(self.entry_gender, 2, 1)
        form_layout.addWidget(create_label("Chức vụ:"), 3, 0)
        form_layout.addWidget(self.entry_position, 3, 1)
        form_layout.addWidget(create_label("Quyền hạn:"), 4, 0)
        form_layout.addWidget(self.entry_role, 4, 1)
        form_layout.addWidget(create_label("Ngày sinh:"), 5, 0)
        form_layout.addWidget(self.entry_dob, 5, 1)
        form_layout.addWidget(create_label("Ngày vào làm:"), 6, 0)
        form_layout.addWidget(self.entry_joindate, 6, 1)

        # Label chữ "Thông báo:"
        self.label_info = QLabel("Thông báo:")
        self.label_info.setStyleSheet(
            "color: #2196F3; font-weight: bold; padding: 4px 8px; font-size: 12px;"
        )
        self.label_info.setMaximumHeight(30)

        # Label để hiển thị nội dung notice động
        self.label_notice = QLabel("")
        self.label_notice.setStyleSheet(
            "color: #2196F3; padding: 4px 8px; background: #e3f2fd; border-radius: 4px; font-size: 11px;"
        )
        self.label_notice.setMaximumHeight(30)
        self.label_notice.setVisible(True)

        # Thêm vào layout, ví dụ đặt label_info ở cột 0, label_notice cột 1 cùng hàng 7
        form_layout.addWidget(self.label_info, 7, 0)
        form_layout.addWidget(self.label_notice, 7, 1)

        # Buttons đơn giản
        button_style = """
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background: #1976D2; }
        """

        btn_capture = QPushButton("Bắt đầu nhận diện")
        btn_capture.setStyleSheet(button_style.replace("#2196F3", "#4CAF50").replace("#1976D2", "#45a049"))
        btn_capture.clicked.connect(self.capture_image)

        btn_stop = QPushButton("Dừng camera")
        btn_stop.setStyleSheet(button_style.replace("#2196F3", "#FF9800").replace("#1976D2", "#F57C00"))
        btn_stop.clicked.connect(self.stop_camera)

        btn_save = QPushButton("Lưu thông tin")
        btn_save.setStyleSheet(button_style)
        btn_save.clicked.connect(self.save_info)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(btn_capture)
        btn_layout.addWidget(btn_stop)
        btn_layout.addWidget(btn_save)

        form_layout.addLayout(btn_layout, 8, 0, 1, 2)

        # Camera preview area - simplified without manual selection
        camera_layout = QVBoxLayout()

        # Live camera preview
        camera_title = QLabel("Camera trực tiếp:")
        camera_title.setStyleSheet("color: #333; font-weight: bold; margin-bottom: 5px;")

        self.camera_preview = QLabel("Camera sẽ hiển thị ở đây")
        self.camera_preview.setFixedSize(320, 240)
        self.camera_preview.setAlignment(Qt.AlignCenter)
        self.camera_preview.setStyleSheet(
            "border: 2px solid #2196F3; border-radius: 8px; background: #f5f5f5; color: #666;")

        # Auto-selected avatar preview
        preview_label = QLabel("Ảnh đại diện đã chọn tự động:")
        preview_label.setStyleSheet("color: #333; font-weight: bold; margin-top: 10px;")

        self.label_image = QLabel("Chưa chọn ảnh")
        self.label_image.setFixedSize(200, 200)
        self.label_image.setAlignment(Qt.AlignCenter)
        self.label_image.setStyleSheet("border: 2px dashed #ccc; border-radius: 8px; color: #666; margin-top: 5px;")

        camera_layout.addWidget(camera_title)
        camera_layout.addWidget(self.camera_preview)
        camera_layout.addWidget(preview_label)
        camera_layout.addWidget(self.label_image)
        camera_layout.addStretch()

        # Main layout với form bên trái, camera bên phải
        main_layout = QHBoxLayout()
        form_widget = QWidget()
        form_widget.setLayout(form_layout)

        camera_widget = QWidget()
        camera_widget.setLayout(camera_layout)

        main_layout.addWidget(form_widget, 2)  # Form chiếm 2/3
        main_layout.addWidget(camera_widget, 1)  # Camera chiếm 1/3

        self.main_layout.addLayout(main_layout)

    def frame_to_base64(self, frame):
        """Convert OpenCV frame to base64 string"""
        import base64

        # Resize frame to reasonable size for storage
        height, width = frame.shape[:2]
        if width > 300:
            ratio = 300.0 / width
            new_width = 300
            new_height = int(height * ratio)
            frame = cv2.resize(frame, (new_width, new_height))

        # Convert to JPEG
        _, buffer = cv2.imencode('.jpg', frame)

        # Convert to base64
        base64_str = base64.b64encode(buffer).decode('utf-8')
        return base64_str

    def calculate_face_quality(self, frame):
        """Calculate face quality score for automatic selection"""
        try:
            # Convert to grayscale for quality analysis
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # Face detection
            face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            if len(faces) == 0:
                return 0

            # Get the largest face
            largest_face = max(faces, key=lambda x: x[2] * x[3])
            x, y, w, h = largest_face

            # Extract face region
            face_roi = gray[y:y + h, x:x + w]

            # Quality metrics
            quality_score = 0

            # 1. Face size (larger is better)
            face_size_score = min(w * h / (frame.shape[0] * frame.shape[1]), 0.3) * 100
            quality_score += face_size_score

            # 2. Sharpness (Laplacian variance)
            laplacian_var = cv2.Laplacian(face_roi, cv2.CV_64F).var()
            sharpness_score = min(laplacian_var / 1000, 1) * 50
            quality_score += sharpness_score

            # 3. Brightness (avoid too dark or too bright)
            brightness = np.mean(face_roi)
            brightness_score = max(0, 50 - abs(brightness - 128) * 0.4)
            quality_score += brightness_score

            # 4. Contrast
            contrast = np.std(face_roi)
            contrast_score = min(contrast / 50, 1) * 30
            quality_score += contrast_score

            return quality_score

        except Exception as e:
            print(f"Error calculating face quality: {e}")
            return 0

    def auto_select_best_avatar(self):
        """Automatically select the best quality avatar from captured frames"""
        try:
            if not self.captured_frames or not self.face_qualities:
                return None

            # Find frame with highest quality score
            best_index = np.argmax(self.face_qualities)
            best_frame = self.captured_frames[best_index]
            best_quality = self.face_qualities[best_index]

            print(f"Auto-selected avatar {best_index} with quality score: {best_quality:.2f}")

            # Update preview
            rgb_frame = cv2.cvtColor(best_frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_frame.shape
            bytes_per_line = ch * w
            qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qt_image).scaled(200, 200, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.label_image.setPixmap(pixmap)

            return best_frame

        except Exception as e:
            print(f"Error in auto_select_best_avatar: {e}")
            return None

    def capture_frames_and_embeddings(self, duration=10):
        """Capture frames và hiển thị trong PyQt widget - with auto avatar selection"""
        try:
            cam = Camera()
            self.current_cam = cam
            cam.capture_for_duration()
            print("Camera mở thành công")
        except Exception as e:
            self.label_notice.setText(f"Lỗi mở camera: {str(e)}")
            self.label_notice.setStyleSheet("color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
            self.label_notice.setVisible(True)
            print(f"Lỗi mở camera: {str(e)}")
            return

        start_time = time.time()
        self.embeddings_list = []
        self.captured_frames = []  # Reset captured frames
        self.face_qualities = []  # Reset face qualities
        print("Đã reset danh sách embeddings, frames và qualities")

        # Tạo timer để update frame
        self.timer = QTimer()
        frame_count = 0

        def update_frame():
            nonlocal frame_count

            try:
                if not cam.cap.isOpened():
                    print("Camera không mở được nữa, dừng timer")
                    self.timer.stop()
                    return

                ret, frame = cam.cap.read()
                if not ret:
                    print("Không đọc được frame, bỏ qua")
                    return

                # Hiển thị frame trong PyQt label
                try:
                    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    h, w, ch = rgb_frame.shape
                    bytes_per_line = ch * w
                    qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                    pixmap = QPixmap.fromImage(qt_image).scaled(320, 240, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.camera_preview.setPixmap(pixmap)
                except Exception as e:
                    print(f"Lỗi cập nhật UI: {e}")

                # Detect face và lưu embedding
                try:
                    _, embedding = self.face_util.detect_face(rgb_frame)

                    if embedding is not None:
                        self.embeddings_list.append(np.array(embedding))

                        # Lưu frame mỗi 30 frame (khoảng 1 giây với 30fps)
                        if frame_count % 30 == 0:
                            print(f"Lưu frame thứ {frame_count}")
                            self.captured_frames.append(frame.copy())

                            # Calculate and store face quality
                            quality = self.calculate_face_quality(frame)
                            self.face_qualities.append(quality)
                            print(f"Frame {len(self.captured_frames) - 1} quality: {quality:.2f}")

                        frame_count += 1
                    else:
                        print("Không phát hiện được embedding trong frame này")
                except Exception as e:
                    print(f"Face detection error: {e}")

                # Lưu frame cuối cùng
                self.current_frame = frame

                # Dừng sau duration
                if time.time() - start_time > duration:
                    print("Đã đủ thời gian, dừng capture")
                    self.timer.stop()

                    # Đảm bào release camera trước khi gọi finish_capture
                    try:
                        cam.release()
                        print("Đã release camera")
                    except Exception as e:
                        print(f"Lỗi release camera: {e}")

                    # Gọi finish_capture với error handling
                    try:
                        self.finish_capture()
                        print("Hoàn thành finish_capture")
                    except Exception as e:
                        print(f"Lỗi trong finish_capture: {e}")
                        # Hiển thị lỗi cho user
                        self.label_notice.setText(f"Lỗi xử lý dữ liệu: {str(e)}")
                        self.label_notice.setStyleSheet(
                            "color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
                        self.label_notice.setVisible(True)

            except Exception as e:
                print(f"Lỗi trong update_frame: {e}")
                # Dừng timer nếu có lỗi nghiêm trọng
                if self.timer and self.timer.isActive():
                    self.timer.stop()
                if cam:
                    try:
                        cam.release()
                    except:
                        pass

        self.timer.timeout.connect(update_frame)
        self.timer.start(33)  # ~30 FPS

    def finish_capture(self):
        """Xử lý sau khi capture xong - with automatic avatar selection"""
        try:
            print(f"Bắt đầu finish_capture, số embeddings: {len(self.embeddings_list)}")

            if len(self.embeddings_list) == 0:
                print("Không có embeddings")
                self.label_notice.setText("Không nhận diện được khuôn mặt. Thử lại!")
                self.label_notice.setStyleSheet(
                    "color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
                self.label_notice.setVisible(True)
                # Reset camera preview
                self.camera_preview.clear()
                self.camera_preview.setText("Camera sẽ hiển thị ở đây")
                return

            print("Tính trung bình embedding...")
            # Tính trung bình embedding
            try:
                avg_embedding = np.mean(np.array(self.embeddings_list), axis=0)
                avg_embedding /= np.linalg.norm(avg_embedding)
                self.current_embedding = avg_embedding
                print(f"Đã tính xong embedding, shape: {avg_embedding.shape}")
            except Exception as e:
                print(f"Lỗi tính embedding: {e}")
                raise e

            print("Tự động chọn avatar tốt nhất...")
            # Automatically select best avatar
            if len(self.captured_frames) > 0:
                try:
                    best_frame = self.auto_select_best_avatar()
                    if best_frame is not None:
                        self.selected_avatar_frame = best_frame
                        print("Đã tự động chọn avatar tốt nhất")
                    else:
                        # Fallback to first frame if auto selection fails
                        self.selected_avatar_frame = self.captured_frames[0]
                        print("Fallback: chọn frame đầu tiên")
                except Exception as e:
                    print(f"Lỗi tự động chọn avatar: {e}")
                    # Fallback to first frame
                    self.selected_avatar_frame = self.captured_frames[0]

            print("Cập nhật thông báo thành công...")
            self.label_notice.setText(
                f"Nhận diện thành công! Đã tự động chọn ảnh đại diện tốt nhất từ {len(self.captured_frames)} ảnh.")
            self.label_notice.setStyleSheet("color: #4CAF50; padding: 8px; background: #e8f5e8; border-radius: 4px;")
            self.label_notice.setVisible(True)

            print("Reset camera preview...")
            # Reset camera preview
            self.camera_preview.clear()
            self.camera_preview.setText("Nhận diện hoàn tất")

            print("finish_capture hoàn thành thành công")

        except Exception as e:
            print(f"Lỗi trong finish_capture: {e}")
            import traceback
            traceback.print_exc()

            # Hiển thị lỗi chi tiết cho user
            self.label_notice.setText(f"Lỗi xử lý dữ liệu: {str(e)}")
            self.label_notice.setStyleSheet("color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
            self.label_notice.setVisible(True)

            # Reset camera preview
            self.camera_preview.clear()
            self.camera_preview.setText("Có lỗi xảy ra")

    def capture_image(self):
        """Bắt đầu capture với preview trong PyQt"""
        try:
            self.label_notice.setText("Đang khởi động camera...")
            self.label_notice.setVisible(True)

            # Reset current data
            self.current_frame = None
            self.current_embedding = None
            self.embeddings_list = []
            self.captured_frames = []
            self.face_qualities = []  # Reset face qualities
            self.selected_avatar_frame = None

            # Clean up any existing camera
            self.cleanup_camera()

            # Reset avatar preview
            self.label_image.clear()
            self.label_image.setText("Chưa chọn ảnh")

            # Bắt đầu capture
            self.capture_frames_and_embeddings(duration=10)

        except Exception as e:
            QMessageBox.critical(self, "Lỗi webcam", str(e))
            self.label_notice.setText("Lỗi khi mở camera!")
            self.label_notice.setStyleSheet("color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
            self.label_notice.setVisible(True)

    def stop_camera(self):
        """Dừng camera thủ công"""
        self.cleanup_camera()
        self.camera_preview.clear()
        self.camera_preview.setText("Camera đã dừng")
        self.label_notice.setText("Camera đã được dừng")
        self.label_notice.setStyleSheet("color: #FF9800; padding: 8px; background: #fff3e0; border-radius: 4px;")
        self.label_notice.setVisible(True)

    def save_info(self):
        name = self.entry_name.text().strip()
        department = self.department.text().strip()
        gender = self.entry_gender.currentText().strip()
        position = self.entry_position.text().strip()
        role = self.entry_role.currentText().strip()
        dob = self.entry_dob.date().toString("yyyy-MM-dd")
        join_date = self.entry_joindate.date().toString("yyyy-MM-dd")

        if not name or not department:
            QMessageBox.warning(self, "Thiếu thông tin", "Vui lòng nhập họ tên và Phòng ban.")
            return

        if self.current_embedding is None:
            QMessageBox.warning(self, "Thiếu dữ liệu nhận diện", "Vui lòng nhận diện khuôn mặt trước.")
            return

        if self.selected_avatar_frame is None:
            QMessageBox.warning(self, "Thiếu ảnh đại diện", "Vui lòng chụp ảnh để hệ thống tự động chọn đại diện.")
            return

        try:
            # Convert selected avatar to base64
            face_img_base64 = self.frame_to_base64(self.selected_avatar_frame)
            face_img_bytes = base64.b64decode(face_img_base64)
            print("face_img_bytes", face_img_bytes)

            # Save employee with face image
            emp_id = self.db.employees.add_employee(
                full_name=name,
                department=department,
                gender=gender,
                position=position,
                dob=dob,
                join_date=join_date,
                face_img=face_img_bytes  # Add face_img parameter
            )

            # Save face encoding
            encoding_str = ','.join(map(str, self.current_embedding.tolist()))
            print("emp_id", emp_id)
            self.db.employees.add_encoding(emp_id, encoding_str)

            # Create user account
            username = str(emp_id)  # Username sử dụng employee ID

            # Set password based on role
            if role == "Admin":
                password = "123"
                db_role = "admin"
            else:  # Nhân viên
                password = "456"
                db_role = "employee"

            # Add user to Users table
            self.db.employees.add_user(
                username=username,
                password=password,
                role=db_role,
                employee_id=emp_id
            )

            self.label_notice.setText(f"Đã lưu thành công: {name} - Tài khoản: {username}")
            self.label_notice.setStyleSheet("color: #4CAF50; padding: 8px; background: #e8f5e8; border-radius: 4px;")
            self.label_notice.setVisible(True)

            QMessageBox.information(self, "Thành công",
                                    f"Đã lưu nhân viên {name} với ảnh đại diện\n"
                                    f"Tài khoản: {username}\n"
                                    f"Mật khẩu: {password}\n"
                                    f"Quyền: {role}")

            # Reset form
            self.reset_form()

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi lưu: {str(e)}")

    def reset_form(self):
        """Reset form after successful save"""
        self.entry_name.clear()
        self.department.clear()
        self.entry_gender.setCurrentIndex(0)
        self.entry_position.clear()
        self.current_frame = None
        self.current_embedding = None
        self.captured_frames = []
        self.selected_avatar_frame = None
        self.label_image.clear()
        self.label_image.setText("Chưa chọn ảnh")
        self.camera_preview.clear()
        self.camera_preview.setText("Camera sẽ hiển thị ở đây")

    def show_register(self):
        try:
            self.cleanup_camera()  # Clean up any running camera
            self.clear_layout(self.main_layout)
            self.page_title.setText("ĐĂNG KÝ NHÂN VIÊN")
            self.build_register_form()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi hiển thị form: {str(e)}")
   # DELETE EMPLOYEES
    def handle_delete_employee(self, row_index):
        """Xử lý khi click nút Delete nhân viên"""
        try:
            employees = self.db.employees.get_all_employees()
            if row_index < len(employees):
                employee = employees[row_index]
                employee_id = employee[0]  # EmployeeID là cột đầu tiên
                employee_name = employee[1]  # FullName là cột thứ hai

                reply = QMessageBox.question(
                    self,
                    "Xác nhận xoá",
                    f"Bạn có chắc muốn xoá nhân viên {employee_name}?",
                    QMessageBox.Yes | QMessageBox.No
                )

                if reply == QMessageBox.Yes:
                    # Gọi db.delete_employee với EmployeeID
                    self.db.employees.delete_employee(employee_id)

                    # Hiển thị thông báo thành công
                    QMessageBox.information(
                        self,
                        "Thành công",
                        f"Đã xoá nhân viên {employee_name} thành công!"
                    )

                    # Refresh lại table
                    self.build_employee_list_ui()

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi xoá: {str(e)}")

    # EDIT EMPLOY

    def load_current_avatar(self):
        """Load và hiển thị ảnh đại diện hiện tại của nhân viên"""
        try:
            employee_id = self.edit_data['employee_id']

            # Get face image from database
            face_img_data = self.db.employees.get_employee_face_image(employee_id)

            if face_img_data:
                # Convert binary data to QPixmap
                pixmap = QPixmap()
                pixmap.loadFromData(face_img_data)

                if not pixmap.isNull():
                    # Scale and display
                    scaled_pixmap = pixmap.scaled(200, 200, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.edit_current_avatar.setPixmap(scaled_pixmap)
                else:
                    self.edit_current_avatar.setText("Lỗi tải ảnh")
            else:
                self.edit_current_avatar.setText("Không có ảnh")

        except Exception as e:
            print(f"Error loading current avatar: {e}")
            self.edit_current_avatar.setText("Lỗi tải ảnh")

    def capture_face_for_edit(self):
        """Bắt đầu capture khuôn mặt cho chế độ edit"""
        try:
            self.edit_label_notice.setText("Đang khởi động camera...")
            self.edit_label_notice.setStyleSheet(
                "color: #2196F3; padding: 8px; background: #e3f2fd; border-radius: 4px;")
            self.edit_label_notice.setVisible(True)

            # Reset edit-specific data
            self.edit_current_frame = None
            self.edit_current_embedding = None
            self.edit_embeddings_list = []
            self.edit_captured_frames = []
            self.edit_face_qualities = []
            self.edit_selected_avatar_frame = None

            # Clean up any existing camera
            self.cleanup_camera()

            # Reset new avatar preview
            self.edit_new_avatar.clear()
            self.edit_new_avatar.setText("Chưa chụp ảnh mới")

            # Bắt đầu capture cho edit mode
            self.capture_frames_for_edit(duration=5)

        except Exception as e:
            QMessageBox.critical(self, "Lỗi webcam", str(e))
            self.edit_label_notice.setText("Lỗi khi mở camera!")
            self.edit_label_notice.setStyleSheet(
                "color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")

    def capture_frames_for_edit(self, duration=5):
        """Capture frames cho chế độ edit"""
        try:
            cam = Camera()
            self.current_cam = cam
            cam.capture_for_duration(seconds=duration)
            print("Camera mở thành công cho edit mode")
        except Exception as e:
            self.edit_label_notice.setText(f"Lỗi mở camera: {str(e)}")
            self.edit_label_notice.setStyleSheet(
                "color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
            return

        start_time = time.time()
        self.edit_embeddings_list = []
        self.edit_captured_frames = []
        self.edit_face_qualities = []
        print("Đã reset danh sách edit embeddings, frames và qualities")

        # Tạo timer để update frame cho edit mode
        self.timer = QTimer()
        frame_count = 0

        def update_edit_frame():
            nonlocal frame_count

            try:
                if not cam.cap.isOpened():
                    print("Camera không mở được nữa, dừng timer")
                    self.timer.stop()
                    return

                ret, frame = cam.cap.read()
                if not ret:
                    print("Không đọc được frame, bỏ qua")
                    return

                # Hiển thị frame trong edit camera preview
                try:
                    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    h, w, ch = rgb_frame.shape
                    bytes_per_line = ch * w
                    qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
                    pixmap = QPixmap.fromImage(qt_image).scaled(320, 240, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.edit_camera_preview.setPixmap(pixmap)
                except Exception as e:
                    print(f"Lỗi cập nhật edit camera UI: {e}")

                # Detect face và lưu embedding cho edit
                try:
                    _, embedding = self.face_util.detect_face(rgb_frame)

                    if embedding is not None:
                        self.edit_embeddings_list.append(np.array(embedding))

                        # Lưu frame mỗi 30 frame
                        if frame_count % 30 == 0:
                            print(f"Lưu edit frame thứ {frame_count}")
                            self.edit_captured_frames.append(frame.copy())

                            # Calculate and store face quality for edit
                            quality = self.calculate_face_quality(frame)
                            self.edit_face_qualities.append(quality)
                            print(f"Edit frame {len(self.edit_captured_frames) - 1} quality: {quality:.2f}")

                        frame_count += 1
                    else:
                        print("Không phát hiện được embedding trong edit frame này")
                except Exception as e:
                    print(f"Edit face detection error: {e}")

                # Lưu frame cuối cùng cho edit
                self.edit_current_frame = frame

                # Dừng sau duration
                if time.time() - start_time > duration:
                    print("Đã đủ thời gian, dừng edit capture")
                    self.timer.stop()

                    try:
                        cam.release()
                        print("Đã release camera cho edit")
                    except Exception as e:
                        print(f"Lỗi release camera cho edit: {e}")

                    try:
                        self.finish_edit_capture()
                        print("Hoàn thành finish_edit_capture")
                    except Exception as e:
                        print(f"Lỗi trong finish_edit_capture: {e}")
                        self.edit_label_notice.setText(f"Lỗi xử lý dữ liệu: {str(e)}")
                        self.edit_label_notice.setStyleSheet(
                            "color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")

            except Exception as e:
                print(f"Lỗi trong update_edit_frame: {e}")
                if self.timer and self.timer.isActive():
                    self.timer.stop()
                if cam:
                    try:
                        cam.release()
                    except:
                        pass

        self.timer.timeout.connect(update_edit_frame)
        self.timer.start(33)  # ~30 FPS

    def finish_edit_capture(self):
        """Xử lý sau khi capture xong cho edit mode"""
        try:
            print(f"Bắt đầu finish_edit_capture, số embeddings: {len(self.edit_embeddings_list)}")

            if len(self.edit_embeddings_list) == 0:
                print("Không có edit embeddings")
                self.edit_label_notice.setText("Không nhận diện được khuôn mặt. Thử lại!")
                self.edit_label_notice.setStyleSheet(
                    "color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
                self.edit_camera_preview.clear()
                self.edit_camera_preview.setText("Camera sẽ hiển thị ở đây")
                return

            print("Tính trung bình edit embedding...")
            # Tính trung bình embedding cho edit
            try:
                avg_embedding = np.mean(np.array(self.edit_embeddings_list), axis=0)
                avg_embedding /= np.linalg.norm(avg_embedding)
                self.edit_current_embedding = avg_embedding
                print(f"Đã tính xong edit embedding, shape: {avg_embedding.shape}")
            except Exception as e:
                print(f"Lỗi tính edit embedding: {e}")
                raise e

            print("Tự động chọn avatar tốt nhất cho edit...")
            # Automatically select best avatar for edit
            if len(self.edit_captured_frames) > 0:
                try:
                    best_frame = self.auto_select_best_edit_avatar()
                    if best_frame is not None:
                        self.edit_selected_avatar_frame = best_frame
                        print("Đã tự động chọn edit avatar tốt nhất")
                    else:
                        self.edit_selected_avatar_frame = self.edit_captured_frames[0]
                        print("Fallback: chọn edit frame đầu tiên")
                except Exception as e:
                    print(f"Lỗi tự động chọn edit avatar: {e}")
                    self.edit_selected_avatar_frame = self.edit_captured_frames[0]

            print("Cập nhật thông báo edit thành công...")
            self.edit_label_notice.setText(
                f"Cập nhật khuôn mặt thành công! Đã chọn ảnh tốt nhất từ {len(self.edit_captured_frames)} ảnh.")
            self.edit_label_notice.setStyleSheet(
                "color: #4CAF50; padding: 8px; background: #e8f5e8; border-radius: 4px;")

            print("Reset edit camera preview...")
            self.edit_camera_preview.clear()
            self.edit_camera_preview.setText("Cập nhật hoàn tất")

            print("finish_edit_capture hoàn thành thành công")

        except Exception as e:
            print(f"Lỗi trong finish_edit_capture: {e}")
            import traceback
            traceback.print_exc()

            self.edit_label_notice.setText(f"Lỗi xử lý dữ liệu: {str(e)}")
            self.edit_label_notice.setStyleSheet(
                "color: #f44336; padding: 8px; background: #ffebee; border-radius: 4px;")
            self.edit_camera_preview.clear()
            self.edit_camera_preview.setText("Có lỗi xảy ra")

    def auto_select_best_edit_avatar(self):
        """Tự động chọn avatar tốt nhất cho edit mode"""
        try:
            if not self.edit_captured_frames or not self.edit_face_qualities:
                return None

            # Find frame with highest quality score
            best_index = np.argmax(self.edit_face_qualities)
            best_frame = self.edit_captured_frames[best_index]
            best_quality = self.edit_face_qualities[best_index]

            print(f"Auto-selected edit avatar {best_index} with quality score: {best_quality:.2f}")

            # Update new avatar preview
            rgb_frame = cv2.cvtColor(best_frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_frame.shape
            bytes_per_line = ch * w
            qt_image = QImage(rgb_frame.data, w, h, bytes_per_line, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qt_image).scaled(200, 200, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.edit_new_avatar.setPixmap(pixmap)

            return best_frame

        except Exception as e:
            print(f"Error in auto_select_best_edit_avatar: {e}")
            return None

    def save_employee_edit(self):
        """Lưu thay đổi employee với khả năng cập nhật khuôn mặt"""
        try:
            # Lấy dữ liệu từ form
            name = self.edit_emp_name.text().strip()
            department = self.edit_emp_department.text().strip()
            gender = self.edit_emp_gender.currentText()

            # Ánh xạ quyền hạn từ giao diện -> DB
            role_display = self.edit_user_role.currentText()
            if role_display == "Admin":
                role = "admin"
            else:
                role = "user"

            position = self.edit_emp_position.text().strip()
            dob = self.edit_emp_dob.date().toString("yyyy-MM-dd")
            join_date = self.edit_emp_joindate.date().toString("yyyy-MM-dd")

            if not name or not department:
                QMessageBox.warning(self, "Thiếu thông tin", "Vui lòng nhập họ tên và Phòng ban.")
                return

            employee_id = self.edit_data['employee_id']

            # Update basic employee info
            self.db.employees.update_employee(employee_id, name, department, gender, position, dob, join_date, role)

            # Update face data nếu có ảnh mới
            if hasattr(self, 'edit_current_embedding') and self.edit_current_embedding is not None:
                print("Cập nhật dữ liệu khuôn mặt mới...")

                # Update face encoding
                encoding_str = ','.join(map(str, self.edit_current_embedding.tolist()))
                self.db.employees.update_face_encoding(employee_id, encoding_str)

                # Update face image nếu có
                if hasattr(self, 'edit_selected_avatar_frame') and self.edit_selected_avatar_frame is not None:
                    face_img_base64 = self.frame_to_base64(self.edit_selected_avatar_frame)
                    face_img_bytes = base64.b64decode(face_img_base64)
                    self.db.employees.update_employee_face_image(employee_id, face_img_bytes)

                QMessageBox.information(self, "Thành công",
                                        f"Đã cập nhật thông tin và khuôn mặt của nhân viên {name}!")
            else:
                QMessageBox.information(self, "Thành công",
                                        f"Đã cập nhật thông tin nhân viên {name}!")

            # Reset edit mode
            self.edit_mode = None
            self.edit_data = None

            # Clean up edit-specific data
            for attr in ['edit_current_embedding', 'edit_embeddings_list',
                         'edit_captured_frames', 'edit_face_qualities',
                         'edit_selected_avatar_frame']:
                if hasattr(self, attr):
                    delattr(self, attr)

            # Quay về danh sách nhân viên
            self.show_employee_list()

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi lưu: {str(e)}")

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget:
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())


    # CLASS ATTENDANCE



    def handle_edit_attendance(self, row_index):
        """Handle edit attendance với original data"""
        try:
            # Sử dụng original_data thay vì displayed data
            original_data = self.attendance_table.original_data
            if row_index < len(original_data):
                selected_record = original_data[row_index]

                # Tạo edit_data với đầy đủ thông tin
                self.edit_data = {
                    'log_id': selected_record[0],  # log_id
                    'employee_id': selected_record[1],  # employee_id
                    'employee_name': selected_record[2],  # full_name
                    'date': selected_record[3],  # work_date
                    'check_in': selected_record[4],  # check_in_str (raw)
                    'check_out': selected_record[5],  # check_out_str (raw)
                    'working_hours': selected_record[6],  # work_hours
                    'status': selected_record[7], # status
                    'note': selected_record[8]  # status
                }

                print(f"Edit data: {self.edit_data}")
                self.show_edit_attendance_form()
            else:
                QMessageBox.warning(self, "Lỗi", "Không thể tìm thấy bản ghi để chỉnh sửa!")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Có lỗi xảy ra khi chỉnh sửa: {str(e)}")
            print(f"Error in handle_edit_attendance: {str(e)}")

    def handle_delete_attendance(self, row_index):
        """Handle delete attendance với original data"""
        try:
            # Sử dụng original_data thay vì displayed data
            original_data = self.attendance_table.original_data
            if row_index < len(original_data):
                selected_record = original_data[row_index]

                log_id = selected_record[0]
                employee_name = selected_record[2]
                work_date = selected_record[3]

                # Xác nhận xóa
                reply = QMessageBox.question(
                    self,
                    "Xác nhận xóa",
                    f"Bạn có chắc chắn muốn xóa bản ghi chấm công?\n\n"
                    f"Nhân viên: {employee_name}\n"
                    f"Ngày: {work_date}",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if reply == QMessageBox.Yes:
                    # Thực hiện xóa
                    success = self.db.attendance.delete_attendance_log(log_id)
                    if success:
                        QMessageBox.information(self, "Thành công", "Đã xóa bản ghi chấm công!")
                        self.build_attendance_stats_ui()  # Refresh lại
                    else:
                        QMessageBox.critical(self, "Lỗi", "Không thể xóa bản ghi!")
            else:
                QMessageBox.warning(self, "Lỗi", "Không thể tìm thấy bản ghi để xóa!")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Có lỗi xảy ra khi xóa: {str(e)}")
            print(f"Error in handle_delete_attendance: {str(e)}")

    def show_edit_attendance_form(self):
        """Hiển thị form sửa thông tin chấm công với logic cải thiện"""
        self.cleanup_camera()
        self.clear_layout(self.main_layout)
        self.page_title.setText("📝 SỬA THÔNG TIN CHẤM CÔNG")

        # Style cho input
        input_style = """
            QLineEdit, QComboBox, QDateEdit, QTimeEdit, QDateTimeEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                font-size: 12px;
                min-width: 250px;
                background-color: #fff;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QTimeEdit:focus, QDateTimeEdit:focus {
                border-color: #2196F3;
            }
        """

        button_style = """
            QPushButton {
                background: #2196F3;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background: #1976D2; }
        """
        cancel_button_style = button_style.replace("#2196F3", "#757575").replace("#1976D2", "#616161")

        def create_label(text):
            lbl = QLabel(text)
            lbl.setStyleSheet("color: #333; font-weight: bold; min-width: 120px;")
            return lbl

        # Thông tin nhân viên
        employee_info = QLabel(f"👤 Nhân viên: {self.edit_data['employee_name']} (ID: {self.edit_data['employee_id']})")
        employee_info.setStyleSheet("color: #555; font-size: 14px; margin-bottom: 12px; font-weight: bold;")

        # Hiển thị thông tin hiện tại
        current_check_in = self.edit_data.get('check_in', 'Chưa check in')
        current_check_out = self.edit_data.get('check_out', 'Chưa check out')

        # Debug - In ra để kiểm tra
        print(f"DEBUG - Raw check_out data: '{current_check_out}'")
        print(f"DEBUG - Raw check_in data: '{current_check_in}'")

        # Parse thời gian hiện tại để hiển thị
        check_in_display = self._format_time_for_display(current_check_in)
        check_out_display = self._format_time_for_display(current_check_out)

        current_info_text = f"""
        📅 Thông tin hiện tại:
        • Ngày: {self.edit_data.get('date', 'N/A')}
        • Check In: {check_in_display}
        • Check Out: {check_out_display}
        • Note:  {self.edit_data.get('note', 'N/A')}
        """

        current_info = QLabel(current_info_text)
        current_info.setStyleSheet("""
            color: #666;
            font-size: 11px;
            background-color: #f5f5f5;
            padding: 10px;
            border-radius: 4px;
            border-left: 3px solid #2196F3;
        """)
        current_info.setWordWrap(True)

        # Date picker
        self.edit_date = QDateEdit()
        self.edit_date.setCalendarPopup(True)
        self.edit_date.setDisplayFormat("dd/MM/yyyy")

        # Parse và set date
        try:
            date_str = self.edit_data['date']
            if isinstance(date_str, str):
                if '/' in date_str:
                    self.edit_date.setDate(QDate.fromString(date_str, "dd/MM/yyyy"))
                elif '-' in date_str:
                    self.edit_date.setDate(QDate.fromString(date_str, "yyyy-MM-dd"))
            else:
                self.edit_date.setDate(QDate.currentDate())
        except:
            self.edit_date.setDate(QDate.currentDate())

        # --- NEW: Check In Time Editor with checkbox ---
        self.check_in_enabled = QCheckBox("Đã check in")
        self.edit_check_in_time = QTimeEdit()
        self.edit_check_in_time.setDisplayFormat("HH:mm:ss")

        # Determine if check-in has a valid value
        has_checkin = self._has_valid_checkin(current_check_in)  # You'll need to implement this helper

        print(f"DEBUG - Has checkin: {has_checkin}")

        self.check_in_enabled.setChecked(has_checkin)
        self.edit_check_in_time.setEnabled(has_checkin)

        # Parse check-in time
        try:
            if has_checkin:
                check_in_time = self._parse_time_from_db(current_check_in)
                self.edit_check_in_time.setTime(check_in_time)
            else:
                self.edit_check_in_time.setTime(QTime(7, 30, 0))  # Default 8:00 AM
        except:
            self.edit_check_in_time.setTime(QTime(7, 30, 0))

        # Connect checkbox to enable/disable check in time
        def toggle_checkin(checked):
            self.edit_check_in_time.setEnabled(checked)
            if not checked:
                self.edit_check_in_time.setTime(QTime(7, 30, 0))  # Reset to default if unchecked

        self.check_in_enabled.toggled.connect(toggle_checkin)
        # --- END NEW: Check In Time Editor with checkbox ---

        # Check Out Time Editor với checkbox để enable/disable
        self.check_out_enabled = QCheckBox("Đã check out")
        self.edit_check_out_time = QTimeEdit()
        self.edit_check_out_time.setDisplayFormat("HH:mm:ss")

        # FIXED: Logic kiểm tra check out được cải thiện
        has_checkout = self._has_valid_checkout(current_check_out)

        print(f"DEBUG - Has checkout: {has_checkout}")

        self.check_out_enabled.setChecked(has_checkout)
        self.edit_check_out_time.setEnabled(has_checkout)

        # Parse check-out time
        try:
            if has_checkout:
                check_out_time = self._parse_time_from_db(current_check_out)
                self.edit_check_out_time.setTime(check_out_time)
            else:
                self.edit_check_out_time.setTime(QTime(17, 0, 0))  # Default 5:00 PM
        except:
            self.edit_check_out_time.setTime(QTime(17, 0, 0))

        # Connect checkbox to enable/disable check out time
        def toggle_checkout(checked):
            self.edit_check_out_time.setEnabled(checked)
            if not checked:
                self.edit_check_out_time.setTime(QTime(17, 0, 0))

        self.check_out_enabled.toggled.connect(toggle_checkout)

        # Apply styles
        for widget in [self.edit_date, self.edit_check_in_time, self.edit_check_out_time]:
            widget.setStyleSheet(input_style)

        # Form layout
        form_layout = QGridLayout()
        form_layout.setSpacing(15)

        form_layout.addWidget(create_label("📅 Ngày làm việc:"), 0, 0)
        form_layout.addWidget(self.edit_date, 0, 1)

        # --- NEW: Check In layout ---
        form_layout.addWidget(create_label("🕐 Giờ check in:"), 1, 0)
        checkin_layout = QVBoxLayout()
        checkin_layout.addWidget(self.check_in_enabled)
        checkin_layout.addWidget(self.edit_check_in_time)
        checkin_widget = QWidget()
        checkin_widget.setLayout(checkin_layout)
        form_layout.addWidget(checkin_widget, 1, 1)
        # --- END NEW: Check In layout ---

        form_layout.addWidget(create_label("🕑 Check out:"), 2, 0)
        checkout_layout = QVBoxLayout()
        checkout_layout.addWidget(self.check_out_enabled)
        checkout_layout.addWidget(self.edit_check_out_time)
        checkout_widget = QWidget()
        checkout_widget.setLayout(checkout_layout)
        form_layout.addWidget(checkout_widget, 2, 1)
        # Note input
        self.note_input = QLineEdit()
        self.note_input.setPlaceholderText("Ghi chú nếu có...")
        self.note_input.setText(self.edit_data.get('note', ''))
        self.note_input.setStyleSheet(input_style)

        form_layout.addWidget(create_label("📝 Ghi chú:"), 3, 0)
        form_layout.addWidget(self.note_input, 3, 1)
        note_text = self.note_input.text().strip()

        form_group = QGroupBox("📄 Thông tin chấm công")
        form_group.setLayout(form_layout)
        form_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                margin-top: 15px;
                padding-top: 15px;
                border: 2px solid #ddd;
                border-radius: 5px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)

        # Buttons
        btn_save = QPushButton("💾 Lưu thay đổi")
        btn_save.setStyleSheet(button_style)
        btn_save.clicked.connect(self.save_attendance_edit)

        btn_cancel = QPushButton("❌ Hủy")
        btn_cancel.setStyleSheet(cancel_button_style)
        btn_cancel.clicked.connect(self.show_attendance_stats)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_cancel)

        # Main layout
        main_form_layout = QVBoxLayout()
        main_form_layout.setSpacing(20)
        main_form_layout.addWidget(employee_info, alignment=Qt.AlignLeft)
        main_form_layout.addWidget(current_info)
        main_form_layout.addWidget(form_group)
        main_form_layout.addLayout(btn_layout)
        main_form_layout.addStretch()

        form_widget = QWidget()
        form_widget.setLayout(main_form_layout)
        self.main_layout.addWidget(form_widget)

    def _has_valid_checkout(self, check_out_value):
        """Kiểm tra xem có check out hợp lệ hay không"""
        if not check_out_value:
            return False

        # Convert to string và strip whitespace
        check_out_str = str(check_out_value).strip()

        # Các trường hợp không hợp lệ
        invalid_values = [
            '',
            'None',
            'null',
            'NULL',
            'Chưa check in',
            'Chưa check out',
            '00:00:00',
            '0:00:00'
        ]

        # Kiểm tra các trường hợp không hợp lệ
        for invalid in invalid_values:
            if check_out_str.lower() == invalid.lower():
                return False

        # Kiểm tra format time hợp lệ
        try:
            parsed_time = self._parse_time_from_db(check_out_str)
            # Kiểm tra không phải là thời gian mặc định (8:00:00)
            if parsed_time.hour() == 8 and parsed_time.minute() == 0 and parsed_time.second() == 0:
                # Nếu giá trị gốc không phải là 8:00:00 thì có thể là lỗi parse
                if '8:00' not in check_out_str and '08:00' not in check_out_str:
                    return True  # Có thể là thời gian hợp lệ khác bị parse sai
            return True
        except:
            return False

    # NEW HELPER FUNCTION
    def _has_valid_checkin(self, check_in_value):
        """Kiểm tra xem có check in hợp lệ hay không"""
        if not check_in_value:
            return False

        check_in_str = str(check_in_value).strip()

        invalid_values = [
            '',
            'None',
            'null',
            'NULL',
            'Chưa check in',
            'Chưa check out',
            '00:00:00',
            '0:00:00'
        ]

        for invalid in invalid_values:
            if check_in_str.lower() == invalid.lower():
                return False

        try:
            parsed_time = self._parse_time_from_db(check_in_str)
            # You might want to refine this if 8:00:00 is a *valid* check-in time for some cases.
            # For now, similar to checkout, we'll consider it valid unless the original string was '8:00' or '08:00'
            if parsed_time.hour() == 8 and parsed_time.minute() == 0 and parsed_time.second() == 0:
                if '8:00' not in check_in_str and '08:00' not in check_in_str:
                    return True
            return True
        except:
            return False

    def _parse_time_from_db(self, time_str):
        """Parse time string from database to QTime - improved version"""
        if not time_str:
            return QTime(7, 30, 0)

        # Convert to string và clean up
        time_str = str(time_str).strip()

        # Các trường hợp không hợp lệ
        if time_str.lower() in ['chưa check in', 'chưa check out', 'none', 'null', '']:
            return QTime(7, 30, 0)

        try:
            # Nếu có microseconds, loại bỏ chúng
            if '.' in time_str:
                time_str = time_str.split('.')[0]

            # Nếu có dạng datetime, chỉ lấy phần time
            if ' ' in time_str:
                parts = time_str.split(' ')
                if len(parts) >= 2:
                    time_str = parts[1]  # Lấy phần thời gian

            # Thử parse với các format khác nhau
            formats = ["HH:mm:ss", "H:mm:ss", "HH:mm", "H:mm"]

            for fmt in formats:
                try:
                    parsed_time = QTime.fromString(time_str, fmt)
                    if parsed_time.isValid():
                        return parsed_time
                except:
                    continue

            # Nếu vẫn không parse được, thử regex
            import re
            time_match = re.match(r'(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?', time_str)
            if time_match:
                hour = int(time_match.group(1))
                minute = int(time_match.group(2))
                second = int(time_match.group(3)) if time_match.group(3) else 0
                return QTime(hour, minute, second)

            # Fallback
            return QTime(7, 30, 0)

        except Exception as e:
            print(f"Error parsing time '{time_str}': {e}")
            return QTime(7, 30, 0)

    def _format_time_for_display(self, time_str):
        """Format time string for display - improved version"""
        if not time_str:
            return 'Chưa có dữ liệu'

        time_str = str(time_str).strip()

        if time_str.lower() in ['chưa check in', 'chưa check out', 'none', 'null', '']:
            return time_str if time_str.lower() in ['chưa check in', 'chưa check out'] else 'Chưa có dữ liệu'

        try:
            parsed_time = self._parse_time_from_db(time_str)
            if parsed_time.isValid():
                return parsed_time.toString("HH:mm:ss")
            else:
                return time_str
        except:
            return time_str

    def save_attendance_edit(self):
        """Lưu thông tin chấm công với logic cải thiện"""
        try:
            # Get form data
            work_date = self.edit_date.date().toString("yyyy-MM-dd")  # Dùng format chuẩn cho database

            # Get check in time (chỉ nếu được enable)
            check_in_str = None
            if self.check_in_enabled.isChecked():
                check_in_time = self.edit_check_in_time.time()
                check_in_str = check_in_time.toString("HH:mm:ss")
            else:
                # If check-in is disabled, set it to None or a specific default for DB
                check_in_str = None  # Or '00:00:00' or 'Chưa check in' depending on your DB
                # Consider how your database handles missing check-in data.
                # If it's a nullable column, None is appropriate.

            # Get check out time (chỉ nếu được enable)
            check_out_str = None
            if self.check_out_enabled.isChecked():
                check_out_time = self.edit_check_out_time.time()
                check_out_str = check_out_time.toString("HH:mm:ss")

                # Validate: check out phải sau check in, only if both are present
                if check_in_str and check_out_str:
                    if QTime.fromString(check_out_str, "HH:mm:ss") <= QTime.fromString(check_in_str, "HH:mm:ss"):
                        QMessageBox.warning(self, "Cảnh báo",
                                            "Giờ check out phải sau giờ check in!")
                        return
                elif check_in_str is None and check_out_str is not None:
                    # If only check-out is set, you might want to add a warning or disallow this
                    QMessageBox.warning(self, "Cảnh báo",
                                        "Không thể có giờ check out nếu chưa có giờ check in!")
                    return

            print(f"Saving attendance:")
            print(f"  SessionID: {self.edit_data['log_id']}")
            print(f"  Employee: {self.edit_data['employee_name']} (ID: {self.edit_data['employee_id']})")
            print(f"  WorkDate: {work_date}")
            print(f"  Check In: {check_in_str}")
            print(f"  Check Out: {check_out_str}")

            # Calculate working hours if both check in and check out are available
            working_hours = 0
            if check_in_str and check_out_str:
                check_in_time_obj = QTime.fromString(check_in_str, "HH:mm:ss")
                check_out_time_obj = QTime.fromString(check_out_str, "HH:mm:ss")

                check_in_minutes = check_in_time_obj.hour() * 60 + check_in_time_obj.minute()
                check_out_minutes = check_out_time_obj.hour() * 60 + check_out_time_obj.minute()
                working_hours = round((check_out_minutes - check_in_minutes) / 60.0, 2)
            elif check_in_str and check_out_str is None:
                # If only check-in is present, working hours might be 0 or calculated based on a default end time
                working_hours = 0
            elif check_in_str is None and check_out_str:
                # If only check-out is present, this is likely an invalid state. Working hours would be 0.
                working_hours = 0
            note_text = self.note_input.text().strip()

            # Update database - thay đổi method call để phù hợp với structure mới
            success = self.db.attendance.update_attendance_log(
                session_id=self.edit_data['log_id'],
                work_date=work_date,
                check_in=check_in_str,
                check_out=check_out_str,
                working_hours=working_hours,
                note=note_text
            )

            if success:
                checkin_info = f"Check In: {check_in_str}" if check_in_str else "Chưa check in"
                checkout_info = f"Check Out: {check_out_str}" if check_out_str else "Chưa check out"
                QMessageBox.information(self, "Thành công",
                                        f"Đã cập nhật thông tin chấm công thành công!\n\n"
                                        f"Nhân viên: {self.edit_data['employee_name']}\n"
                                        f"Ngày: {work_date}\n"
                                        f"{checkin_info}\n"
                                        f"{checkout_info}\n"
                                        f"Giờ làm việc: {working_hours} giờ")

                # Return to attendance stats
                self.show_attendance_stats()
            else:
                QMessageBox.critical(self, "Lỗi", "Không thể cập nhật thông tin chấm công!")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Có lỗi xảy ra khi lưu thông tin: {str(e)}")
            print(f"Error saving attendance: {str(e)}")

    def build_attendance_stats_ui(self):
        try:
            print("[INFO] Bắt đầu dựng lại giao diện thống kê")
            self.cleanup_camera()
            self.clear_layout(self.main_layout)

            print("[INFO] Lấy dữ liệu chấm công")
            self.original_attendance_data = self.db.attendance.get_attendance_logs()
            self.filtered_attendance_data = self.original_attendance_data.copy()

            print("[INFO] Tạo layout thống kê")
            self.stats_layout = QHBoxLayout()
            self.update_stats_display()

            print("[INFO] Tạo nút làm mới")
            refresh_component = CustomButton("🔄 Làm mới", button_type="primary")
            refresh_component.clicked.connect(self.build_attendance_stats_ui)

            print("[INFO] Tạo bộ lọc")
            filter_widgets = self._create_filter_widgets()
            filter_layout = self._create_filter_layout(filter_widgets)
            filter_frame = self._create_filter_frame(filter_layout)

            print("[INFO] Tạo bảng chấm công")
            self.create_attendance_table()
            self.apply_realtime_filter()

            print("[INFO] Lắp ráp UI")
            self.main_layout.addLayout(self.stats_layout)
            self.main_layout.addLayout(refresh_component.create_button_layout())
            self.main_layout.addWidget(filter_frame)
            self.main_layout.addWidget(self.attendance_table)

            print("[INFO] Hoàn tất giao diện")
        except Exception as e:
            print(f"[LỖI] Xảy ra lỗi: {str(e)}")
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi: {str(e)}")

    def get_yellow_button_style(self):
        """Yellow button style for clear filter"""
        return """
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
                font-weight: 500;
                font-family: 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #e67e22;
            }
            QPushButton:pressed {
                background-color: #d35400;
            }
        """

    def get_green_button_style(self):
        """Green button style for export"""
        return """
            QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 12px;
                font-weight: 500;
                font-family: 'Segoe UI';
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
            QPushButton:pressed {
                background-color: #229954;
            }
        """
    def _create_filter_widgets(self):
        """Create all filter widgets and return as dict"""
        widgets = {}

        # Date widgets
        widgets['from_date'] = self._create_date_edit(-30, self.apply_realtime_filter)
        widgets['to_date'] = self._create_date_edit(0, self.apply_realtime_filter)

        # Search widget
        widgets['employee_search'] = self._create_line_edit("Mã hoặc tên...", 160, self.apply_realtime_filter)

        # Combo widgets
        widgets['status_combo'] = self._create_combo(
            ["Tất cả", "Đúng giờ", "Đi trễ", "Vắng", "Đi trễ về sớm"], 100, self.apply_realtime_filter)
        widgets['hours_combo'] = self._create_combo(
            ["Tất cả", ">= 8h", "< 8h", ">= 4h", "< 4h"], 100, self.apply_realtime_filter)

        # Action buttons
        widgets['clear_btn'] = self._create_button("🗑️ Xóa bộ lọc", 120,
                                                   self.get_compact_clear_button_style(), self.clear_all_filters)
        widgets['export_btn'] = self._create_button("📊 Xuất Excel", 110,
                                                    self.get_compact_export_button_style(), self.export_filtered_data)

        # Info label
        widgets['info_label'] = QLabel("Hiển thị tất cả bản ghi")
        widgets['info_label'].setStyleSheet("""
            QLabel {
                color: #27ae60; font-weight: 600; font-size: 12px; font-family: 'Segoe UI';
                background: #ecf8f0; padding: 6px 12px; border-radius: 15px;
                border: 1px solid #a8e6c8;
            }
        """)

        # Store references for later use
        self.employee_search = widgets['employee_search']
        self.from_date_edit = widgets['from_date']
        self.to_date_edit = widgets['to_date']
        self.status_combo = widgets['status_combo']
        self.hours_combo = widgets['hours_combo']
        self.filter_info_label = widgets['info_label']

        return widgets

    def _create_filter_layout(self, widgets):
        """Create the complete filter layout with full width distribution"""
        # Main filter row - distribute evenly across full width
        main_row = QHBoxLayout()
        main_row.setSpacing(0)  # Remove spacing, use stretch instead
        main_row.setContentsMargins(0, 0, 0, 0)

        # Date group
        date_group = self._create_labeled_group([
            ("Từ:", widgets['from_date']),
            ("Đến:", widgets['to_date'])
        ])

        # Search group
        search_group = self._create_labeled_group([
            ("Nhân viên:", widgets['employee_search'])
        ])

        # Status group
        status_group = self._create_labeled_group([
            ("Trạng thái:", widgets['status_combo']),
            ("Giờ làm:", widgets['hours_combo'])
        ])

        # Add groups with equal stretch factors to distribute full width
        main_row.addLayout(date_group, 3)  # 3 parts for date group
        main_row.addLayout(search_group, 2)  # 2 parts for search group
        main_row.addLayout(status_group, 3)  # 3 parts for status group
        # No addStretch() - let groups fill the entire width

        # Action row
        action_row = QHBoxLayout()
        action_row.setSpacing(15)  # Tăng khoảng cách giữa các nút
        action_row.setContentsMargins(8, 0, 8, 0)  # Thêm margins trái phải
        action_row.addWidget(widgets['clear_btn'])
        action_row.addWidget(widgets['export_btn'])
        action_row.addStretch()  # Push info label to the right
        action_row.addWidget(widgets['info_label'])

        # Combine all
        layout = QVBoxLayout()
        layout.setSpacing(18)  # Tăng khoảng cách trên dưới
        layout.setContentsMargins(10, 15, 10, 15)  # Thêm margins trái phải trên dưới
        layout.addLayout(main_row)
        layout.addLayout(action_row)

        return layout

    def _create_filter_frame(self, layout):
        """Version with white background for title and full width"""
        group_box = QGroupBox("🔍 Bộ lọc tìm kiếm")
        group_box.setLayout(layout)

        # Set size policy to expand horizontally
        group_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        group_box.setStyleSheet("""
                QGroupBox {
                    font-weight: bold;
                    font-size: 14px;
                    color: #2c3e50;
                    font-family: 'Segoe UI';
                    border: 2px solid #7f8c8d;
                    border-radius: 8px;
                    background-color: #f8f9fa;
                    margin-top: 1ex;
                    padding-top: 10px;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    padding: 2px 12px;
                    background-color: white;
                    color: #2c3e50;
                    font-weight: bold;
                    left: 10px;
                    border: none;
                }
            """)

        group_box.setContentsMargins(20, 20, 20, 20)  # Tăng margins của GroupBox
        return group_box

    def _create_labeled_group(self, items):
        """Create horizontal group with labels and widgets that expand to fill space"""
        group = QHBoxLayout()
        group.setSpacing(8)
        group.setContentsMargins(8, 5, 8, 5)  # Tăng margins cho group

        for label_text, widget in items:
            label = QLabel(label_text)
            label.setStyleSheet(self.get_compact_label_style())
            label.setMinimumWidth(60)
            label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)  # Right align labels

            # Set widget to expand
            widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

            group.addWidget(label)
            group.addWidget(widget, 1)  # Give widget stretch factor

        return group

    def _create_date_edit(self, days_offset, callback):
        """Create configured date edit widget that expands"""
        date_edit = QDateEdit()
        date_edit.setDate(QDate.currentDate().addDays(days_offset))
        date_edit.setDisplayFormat("dd/MM/yyyy")
        date_edit.setCalendarPopup(True)
        date_edit.setMinimumWidth(130)
        date_edit.setFixedHeight(32)
        date_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        date_edit.setStyleSheet(self.get_compact_date_edit_style())
        date_edit.dateChanged.connect(callback)
        return date_edit

    def _create_line_edit(self, placeholder, width, callback):
        """Create configured line edit widget that expands"""
        line_edit = QLineEdit()
        line_edit.setPlaceholderText(placeholder)
        line_edit.setMinimumWidth(width)
        line_edit.setFixedHeight(32)
        line_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        line_edit.setStyleSheet(self.get_compact_line_edit_style())
        line_edit.textChanged.connect(callback)
        return line_edit
    def _create_combo_box(self, items, width, callback):
        """Create configured combo box widget that expands"""
        combo_box = QComboBox()
        combo_box.addItems(items)
        combo_box.setMinimumWidth(width)
        combo_box.setFixedHeight(32)
        combo_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        combo_box.setStyleSheet(self.get_compact_combo_box_style())
        combo_box.currentTextChanged.connect(callback)
        return combo_box

    def _create_combo_box(self, items, width, callback):
        """Create configured combo box widget with flexible width"""
        combo_box = QComboBox()
        combo_box.addItems(items)
        combo_box.setMinimumWidth(width)  # Use minimum width instead of fixed
        combo_box.setFixedHeight(32)
        combo_box.setStyleSheet(self.get_compact_combo_box_style())
        combo_box.currentTextChanged.connect(callback)
        return combo_box

    def _create_combo(self, items, width, callback):
        """Create configured combo box"""
        combo = QComboBox()
        combo.addItems(items)
        combo.setFixedSize(width, 32)
        combo.setStyleSheet(self.get_compact_combo_style())
        combo.currentTextChanged.connect(callback)
        return combo

    def _create_button(self, text, width, style, callback):
        """Create configured button"""
        button = QPushButton(text)
        button.setFixedSize(width, 32)
        button.setStyleSheet(style)
        button.clicked.connect(callback)
        return button

    # Compact style methods (consolidated)
    def get_compact_label_style(self):
        return "QLabel { color: #495057; font-weight: 600; font-size: 12px; font-family: 'Segoe UI'; min-width: 60px; }"

    def get_compact_date_edit_style(self):
        return """
            QDateEdit { border: 1px solid #ced4da; border-radius: 6px; padding: 4px 8px; 
                background: white; font-size: 11px; font-family: 'Segoe UI'; }
            QDateEdit:focus { border-color: #007bff; box-shadow: 0 0 0 2px rgba(0,123,255,.25); }
            QDateEdit::drop-down { subcontrol-origin: padding; subcontrol-position: top right; 
                width: 20px; border-left: 1px solid #ced4da; background: #f8f9fa; }
        """

    def get_compact_line_edit_style(self):
        return """
            QLineEdit { border: 1px solid #ced4da; border-radius: 6px; padding: 6px 10px; 
                background: white; font-size: 11px; font-family: 'Segoe UI'; }
            QLineEdit:focus { border-color: #007bff; box-shadow: 0 0 0 2px rgba(0,123,255,.25); }
            QLineEdit::placeholder { color: #6c757d; font-style: italic; }
        """

    def get_compact_combo_style(self):
        return """
            QComboBox { border: 1px solid #ced4da; border-radius: 6px; padding: 4px 8px; 
                background: white; font-size: 11px; font-family: 'Segoe UI'; }
            QComboBox:focus { border-color: #007bff; }
            QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; 
                width: 20px; border-left: 1px solid #ced4da; background: #f8f9fa; }
            QComboBox QAbstractItemView { border: 1px solid #ced4da; 
                selection-background-color: #007bff; background: white; }
        """

    def get_compact_clear_button_style(self):
        return """
            QPushButton { background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #fff3cd, stop:1 #ffeaa7);
                border: 1px solid #f39c12; border-radius: 6px; color: #d68910; width: 120px; height: 35px;
                font-weight: 600; font-size: 11px; font-family: 'Segoe UI'; }
            QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ffeaa7, stop:1 #fdcb6e);
                border-color: #e67e22; }
            QPushButton:pressed { background: #fdcb6e; border-color: #d35400; }
        """

    def get_compact_export_button_style(self):
        return """
            QPushButton { background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #d1f2eb, stop:1 #a8e6c8);
                border: 1px solid #27ae60; border-radius: 6px; color: #1e8449; width: 110px; height: 35px;
                font-weight: 600; font-size: 11px; font-family: 'Segoe UI'; }
            QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #a8e6c8, stop:1 #7bdcb5);
                border-color: #229954; }
            QPushButton:pressed { background: #7bdcb5; border-color: #1e8449; }
        """

    def apply_realtime_filter(self):
        """Apply all filter conditions in real-time"""
        try:
            # Start with original data
            filtered_data = self.original_attendance_data.copy()

            # Apply date filter
            from_date = self.from_date_edit.date().toPyDate()
            to_date = self.to_date_edit.date().toPyDate()
            filtered_data = self.filter_data_by_date(filtered_data, from_date, to_date)

            # Apply employee search filter
            employee_text = self.employee_search.text().strip().lower()
            if employee_text:
                filtered_data = [
                    record for record in filtered_data
                    if employee_text in str(record[1]).lower() or  # employee_id
                       employee_text in str(record[2]).lower()  # full_name
                ]

            # Apply status filter
            status_filter = self.status_combo.currentText()
            if status_filter != "Tất cả":
                filtered_data = [
                    record for record in filtered_data
                    if record[7] == status_filter  # status column
                ]

            # Apply work hours filter
            hours_filter = self.hours_combo.currentText()
            if hours_filter != "Tất cả":
                filtered_data = self.filter_by_work_hours(filtered_data, hours_filter)

            # Update filtered data
            self.filtered_attendance_data = filtered_data

            # Update display
            self.update_stats_display()
            self.update_attendance_table(filtered_data)
            self.update_filter_info()

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi áp dụng bộ lọc: {str(e)}")

    def filter_by_work_hours(self, data, hours_condition):
        """Lọc theo điều kiện giờ làm"""
        import re
        filtered = []
        cond = hours_condition.replace(" ", "").lower()

        for record in data:
            try:
                match = re.search(r'(\d+(?:\.\d+)?)', str(record[6]))
                if not match:
                    continue
                hours = float(match.group(1))

                if ">=" in cond and "8" in cond and hours >= 8:
                    filtered.append(record)
                elif "<" in cond and "8" in cond and hours < 8:
                    filtered.append(record)
                elif ">=" in cond and "4" in cond and hours >= 4:
                    filtered.append(record)
                elif "<" in cond and "4" in cond and hours < 4:
                    filtered.append(record)
            except:
                continue
        return filtered

    def update_stats_display(self):
        """Update statistics cards based on filtered data"""
        # ✅ Xóa các widget thống kê cũ
        self.clear_layout(self.stats_layout)

        # ✅ Tính toán lại số liệu
        present_count = len(
            [r for r in self.filtered_attendance_data if r[7] in ['Đúng giờ', 'Đi trễ', 'Đi trễ về sớm', 'Về sớm']])
        right_count = len([r for r in self.filtered_attendance_data if r[7] == 'Đúng giờ'])
        late_count = len([r for r in self.filtered_attendance_data if r[7] == 'Đi trễ'])
        early_leave_count = len([r for r in self.filtered_attendance_data if r[7] == 'Về sớm'])
        late_early_count = len([r for r in self.filtered_attendance_data if r[7] == 'Đi trễ về sớm'])
        absent_count = len([r for r in self.filtered_attendance_data if r[7] == 'Vắng'])

        # ✅ Hiển thị các thẻ thống kê mới (stat cards)
        self.stats_layout.addWidget(self.create_stat_card("Có mặt", str(present_count), "#4CAF50"))
        self.stats_layout.addWidget(self.create_stat_card("Đúng giờ", str(right_count), "#4CAF50"))
        self.stats_layout.addWidget(self.create_stat_card("Đi trễ", str(late_count), "#FF9800"))
        self.stats_layout.addWidget(self.create_stat_card("Về sớm", str(early_leave_count), "#FF5722"))
        self.stats_layout.addWidget(self.create_stat_card("Trễ & Sớm", str(late_early_count), "#FF7043"))
        self.stats_layout.addWidget(self.create_stat_card("Vắng", str(absent_count), "#F44336"))

        self.stats_layout.addStretch()

    def update_filter_info(self):
        """Cập nhật nhãn thông tin bộ lọc"""
        total_original = len(self.original_attendance_data)
        total_filtered = len(self.filtered_attendance_data)

        # Luôn hiển thị số lượng bản ghi đã lọc trên tổng số bản ghi
        self.filter_info_label.setText(f"Hiển thị {total_filtered}/{total_original} bản ghi")

        self.filter_info_label.setStyleSheet("""
            QLabel {
                color: #1976D2;
                font-weight: 600;
                font-size: 13px;
                font-family: 'Segoe UI';
                text-shadow: 0px 1px 1px rgba(0,0,0,0.1);
            }
        """)

    def clear_all_filters(self):
        """Clear all filter conditions"""
        # Disconnect signals temporarily to avoid multiple triggers
        self.from_date_edit.dateChanged.disconnect()
        self.to_date_edit.dateChanged.disconnect()
        self.employee_search.textChanged.disconnect()
        self.status_combo.currentTextChanged.disconnect()
        self.hours_combo.currentTextChanged.disconnect()

        # Reset all filters
        self.from_date_edit.setDate(QDate.currentDate().addDays(-30))
        self.to_date_edit.setDate(QDate.currentDate())
        self.employee_search.clear()
        self.status_combo.setCurrentText("Tất cả")
        self.hours_combo.setCurrentText("Tất cả")

        # Reconnect signals
        self.from_date_edit.dateChanged.connect(self.apply_realtime_filter)
        self.to_date_edit.dateChanged.connect(self.apply_realtime_filter)
        self.employee_search.textChanged.connect(self.apply_realtime_filter)
        self.status_combo.currentTextChanged.connect(self.apply_realtime_filter)
        self.hours_combo.currentTextChanged.connect(self.apply_realtime_filter)

        # Apply filter
        self.apply_realtime_filter()

    def export_filtered_data(self):
        """Export filtered data to Excel"""
        if not self.filtered_attendance_data:
            QMessageBox.information(self, "Thông báo", "Không có dữ liệu để xuất")
            return

        try:
            from PyQt5.QtWidgets import QFileDialog
            import pandas as pd
            from datetime import datetime

            # Choose save location
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Lưu file Excel",
                f"DuLieuChamCong_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if file_path:
                # Convert data to DataFrame
                df_data = []
                for record in self.filtered_attendance_data:
                    df_data.append({
                        'Mã nhân viên': record[1],
                        'Họ tên': record[2],
                        'Ngày làm việc': record[3],
                        'Giờ vào': self._format_time_display(record[4]),
                        'Giờ ra': self._format_time_display(record[5]),
                        'Tổng giờ làm': record[6],
                        'Trạng thái': record[7],
                        'Ghi chú': record[8]
                    })

                df = pd.DataFrame(df_data)
                df.to_excel(file_path, index=False)

                QMessageBox.information(self, "Thành công", f"Đã xuất {len(df_data)} bản ghi ra file Excel!")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi xuất file: {str(e)}")

    def create_attendance_table(self):
        """Create attendance table with initial data"""
        attendance_headers = [
            "Mã nhân viên", "Họ tên", "Ngày làm việc", "Giờ vào",
            "Giờ ra", "Tổng giờ làm", "Trạng thái", "Ghi chú"
        ]

        # Format display data
        display_data = []
        for record in self.filtered_attendance_data:
            formatted_record = [
                record[1],  # employee_id
                record[2],  # full_name
                record[3],  # work_date
                self._format_time_display(record[4]),  # check_in_str
                self._format_time_display(record[5]),  # check_out_str
                record[6],  # work_hours
                record[7],  # status
                record[8]  # note
            ]
            display_data.append(formatted_record)

        self.attendance_table = CustomTableWidget(
            headers=attendance_headers,
            data=display_data,
            show_detail_btn=True
        )

        # Store original data for edit/delete operations
        self.attendance_table.original_data = self.filtered_attendance_data

        # Connect signals
        self.attendance_table.edit_clicked.connect(self.handle_edit_attendance)
        self.attendance_table.delete_clicked.connect(self.handle_delete_attendance)
        self.attendance_table.detail_clicked.connect(self.handle_view_attendance_details)

    def update_attendance_table(self, filtered_data):
        """Update table with filtered data"""
        # Format display data
        display_data = []
        for record in filtered_data:
            formatted_record = [
                record[1],  # employee_id
                record[2],  # full_name
                record[3],  # work_date
                self._format_time_display(record[4]),  # check_in_str
                self._format_time_display(record[5]),  # check_out_str
                record[6],  # work_hours
                record[7],  # status
                record[8]  # note
            ]
            display_data.append(formatted_record)

        # Update table data
        self.attendance_table.refresh_data(display_data)
        self.attendance_table.original_data = filtered_data

    # Style helper methods

    def filter_attendance_by_date(self):
        from_date = self.from_date_edit.date().toPyDate()
        to_date = self.to_date_edit.date().toPyDate()

        try:
            attendance_data = self.db.attendance.get_attendance_logs()
            # Áp dụng bộ lọc ngày
            filtered_data = self.filter_data_by_date(attendance_data, from_date, to_date)

            # Cập nhật bảng với dữ liệu đã lọc
            self.update_attendance_table(filtered_data)

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi truy vấn hoặc lọc dữ liệu: {str(e)}")
    def filter_data_by_date(self, data, from_date, to_date):
        """Filter attendance data by date range"""
        filtered_data = []

        for record in data:
            try:
                # record[3] is work_date
                work_date_str = record[3]

                # Parse date from different formats
                if isinstance(work_date_str, str):
                    # Try different date formats
                    date_formats = ['%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']
                    work_date = None

                    for fmt in date_formats:
                        try:
                            work_date = datetime.strptime(work_date_str, fmt).date()
                            break
                        except ValueError:
                            continue

                    if work_date is None:
                        continue  # Skip records with unparseable dates
                else:
                    work_date = work_date_str if hasattr(work_date_str, 'year') else None

                if work_date and from_date <= work_date <= to_date:
                    filtered_data.append(record)

            except Exception as e:
                print(f"Error filtering record {record}: {e}")
                continue

        return filtered_data

    def clear_date_filter(self):
        """Clear date filter and show all data"""
        # Reset date fields to default
        self.from_date_edit.setDate(QDate.currentDate().addDays(-30))
        self.to_date_edit.setDate(QDate.currentDate())

        # Reload all data
        self.filter_attendance_by_date()

    def update_stat_cards(self, present_count, late_count, absent_count):
        """Update the statistics cards with new values"""
        # This method assumes you have stored references to the stat cards
        # You may need to modify create_stat_card to return a reference
        # or rebuild the stats section entirely

        # For now, we'll rebuild the attendance stats UI
        # In a production app, you'd want to update just the numbers
        pass  # Implement based on your create_stat_card method
    def handle_view_attendance_details(self, original_row_index):
        """
        Handles the 'Xem chi tiết' button click.
        Extracts employee_id and work_date from the original data and queries the database.
        """
        print(f"[DEBUG] original_row_index: {original_row_index}")

        if 0 <= original_row_index < len(self.attendance_table.original_data):
            selected_record = self.attendance_table.original_data[original_row_index]
            print(f"[DEBUG] selected_record: {selected_record}")

            employee_id = selected_record[1]
            work_date = selected_record[3]
            print(f"[DEBUG] employee_id: {employee_id}")
            print(f"[DEBUG] work_date: {work_date}")

            try:
                # Lấy chi tiết chấm công
                detail_logs = self.db.attendance.get_attendance_logs_detail(employee_id, work_date)
                print(f"[DEBUG] detail_logs: {detail_logs}")
            except Exception as e:
                print(f"[ERROR] Lỗi khi gọi get_attendance_logs_detail: {str(e)}")
                QMessageBox.critical(self, "Lỗi", f"Lỗi khi lấy chi tiết chấm công:\n{str(e)}")
                return

            if detail_logs:
                try:
                    # Gọi Dialog chi tiết
                    detail_dialog = AttendanceDetailDialog(employee_id, work_date, detail_logs, parent=self)
                    detail_dialog.exec_()
                except Exception as e:
                    print(f"[ERROR] Lỗi khi mở dialog: {str(e)}")
                    QMessageBox.critical(self, "Lỗi", f"Lỗi khi hiển thị chi tiết:\n{str(e)}")
            else:
                print("[DEBUG] Không có log chi tiết")
                QMessageBox.information(self, "Thông báo",
                                        f"Không tìm thấy chi tiết chấm công cho {employee_id} ngày {work_date}.")
        else:
            print("[ERROR] original_row_index nằm ngoài giới hạn")
            QMessageBox.warning(self, "Lỗi", "Không tìm thấy dữ liệu cho hàng được chọn.")

    def _format_time_display(self, time_str):
        """Format time string để chỉ hiển thị HH:MM:SS"""
        if not time_str:
            return "Chưa check"

        time_str = str(time_str).strip()

        # Các trường hợp không hợp lệ
        if time_str.lower() in ['chưa check in', 'chưa check out', 'none', 'null', '']:
            return "Chưa check"

        try:
            # Nếu có microseconds (.0000000), cắt bỏ
            if '.' in time_str:
                time_str = time_str.split('.')[0]

            # Nếu có datetime format (YYYY-MM-DD HH:MM:SS), chỉ lấy phần time
            if ' ' in time_str and len(time_str.split(' ')) >= 2:
                time_str = time_str.split(' ')[1]

            # Validate format HH:MM:SS
            time_parts = time_str.split(':')
            if len(time_parts) >= 2:
                hour = int(time_parts[0])
                minute = int(time_parts[1])
                second = int(time_parts[2]) if len(time_parts) > 2 else 0

                # Format lại thành HH:MM:SS
                return f"{hour:02d}:{minute:02d}:{second:02d}"

            return time_str

        except Exception as e:
            print(f"Error formatting time '{time_str}': {e}")
            return str(time_str)

    def update_attendance_table(self, attendance_data):
        """Cập nhật bảng chấm công và thống kê"""

        # ✅ Thống kê số lượng theo trạng thái
        present_count = len([
            record for record in attendance_data
            if record[7] in ['Đúng giờ', 'Đi trễ', 'Về sớm', 'Đi trễ về sớm']
        ])
        right_count = len([record for record in attendance_data if record[7] == 'Đúng giờ'])
        late_count = len([record for record in attendance_data if record[7] == 'Đi trễ'])
        early_leave_count = len([record for record in attendance_data if record[7] == 'Về sớm'])
        late_early_count = len([record for record in attendance_data if record[7] == 'Đi trễ về sớm'])
        absent_count = len([record for record in attendance_data if record[7] == 'Vắng'])

        # ✅ Cập nhật các QLabel thống kê (nếu có)
        # self.present_label.setText(str(present_count))
        # self.right_label.setText(str(right_count))
        # self.late_label.setText(str(late_count))
        # self.early_label.setText(str(early_leave_count))
        # self.late_early_label.setText(str(late_early_count))
        # self.absent_label.setText(str(absent_count))

        # ✅ Định dạng lại dữ liệu để hiển thị trong bảng
        display_data = []
        for record in attendance_data:
            formatted_record = [
                record[1],  # employee_id
                record[2],  # full_name
                record[3],  # work_date
                self._format_time_display(record[4]),  # check_in
                self._format_time_display(record[5]),  # check_out
                record[6],  # work_hours
                record[7],  # status
                record[8]  # note
            ]
            display_data.append(formatted_record)

        # ✅ Cập nhật bảng
        self.attendance_table.refresh_data(display_data)
        self.attendance_table.original_data = attendance_data

    def handle_delete_attendance(self, row_index):
        """Fixed delete method for new database structure"""
        try:
            attendance_data = self.db.attendance.get_attendance_logs()
            if row_index < len(attendance_data):
                attendance_record = attendance_data[row_index]

                # Get info from record (based on WorkSessions structure)
                session_id = attendance_record[0]  # SessionID
                employee_name = attendance_record[2]  # FullName
                attendance_date = attendance_record[3]  # WorkDate
                check_in_time = attendance_record[4]  # CheckIn
                status = attendance_record[7]  # Status

                message = (f"Bạn có chắc muốn xoá bản ghi chấm công?\n\n"
                           f"Nhân viên: {employee_name}\n"
                           f"Ngày: {attendance_date}\n"
                           f"Giờ vào: {check_in_time}\n"
                           f"Trạng thái: {status}")

                reply = QMessageBox.question(
                    self,
                    "Xác nhận xoá",
                    message,
                    QMessageBox.Yes | QMessageBox.No
                )

                if reply == QMessageBox.Yes:
                    # Delete by SessionID
                    success = self.db.attendance.delete_attendance_log(session_id)

                    if success:
                        QMessageBox.information(
                            self,
                            "Thành công",
                            f"Đã xoá bản ghi chấm công của {employee_name} vào ngày {attendance_date}!"
                        )
                        # Refresh table
                        self.build_attendance_stats_ui()
                    else:
                        QMessageBox.critical(self, "Lỗi", "Không thể xoá bản ghi!")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi xoá: {str(e)}")

    def show_attendance_stats(self):
        self.clear_layout(self.main_layout)
        self.page_title.setText("THỐNG KÊ CHẤM CÔNG")
        self.build_attendance_stats_ui()

    def closeEvent(self, event):
        """Handle application close event"""
        self.cleanup_camera()
        event.accept()

class AttendanceDetailDialog(QDialog):
    def __init__(self, employee_id, work_date, logs, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Chi tiết chấm công - {employee_id}")

        # Kích thước tối ưu
        self.resize(900, 650)
        self.setMinimumSize(800, 600)

        # Store logs for export
        self.logs = logs

        # Set dialog properties
        self.setModal(True)
        self.setAttribute(Qt.WA_DeleteOnClose)

        # Giao diện đơn giản và sạch sẽ
        self.setStyleSheet("""
            QDialog {
                background-color: #ffffff;
                font-family: 'Segoe UI', Arial, sans-serif;
            }

            QLabel#titleLabel {
                background-color: #2c3e50;
                color: white;
                padding: 15px;
                border-radius: 8px;
                font-size: 18px;
                font-weight: bold;
            }

            QLabel#infoLabel {
                background-color: #f8f9fa;
                color: #495057;
                padding: 10px;
                border-radius: 6px;
                border-left: 4px solid #007bff;
                font-size: 12px;
                margin: 3px;
            }

            QFrame.statsCard {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 15px;
                margin: 5px;
            }

            QLabel.statTitle {
                color: #6c757d;
                font-size: 11px;
                font-weight: 500;
                margin-bottom: 5px;
            }

            QLabel.statValue {
                color: #212529;
                font-size: 18px;
                font-weight: bold;
            }

            QTableWidget {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                gridline-color: #dee2e6;
                font-size: 12px;
            }

            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #f8f9fa;
            }

            QTableWidget::item:selected {
                background-color: #007bff;
                color: white;
            }

            QHeaderView::section {
                background-color: #f8f9fa;
                color: #495057;
                padding: 12px;
                border: 1px solid #dee2e6;
                font-weight: 600;
                font-size: 11px;
            }

            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-size: 12px;
                font-weight: 500;
                min-width: 80px;
            }

            QPushButton:hover {
                background-color: #0056b3;
            }

            QPushButton:pressed {
                background-color: #004085;
            }

            QPushButton:disabled {
                background-color: #6c757d;
                color: #ffffff;
            }

            QPushButton.export {
                background-color: #28a745;
            }

            QPushButton.export:hover {
                background-color: #218838;
            }

            QPushButton.close {
                background-color: #dc3545;
            }

            QPushButton.close:hover {
                background-color: #c82333;
            }

            QPushButton.view {
                background-color: #6f42c1;
                padding: 6px 12px;
                font-size: 11px;
                min-width: 60px;
            }

            QPushButton.view:hover {
                background-color: #5a32a3;
            }
        """)

        self.setupUI(employee_id, work_date, logs)

    def setupUI(self, employee_id, work_date, logs):
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header
        header_widget = self.createHeaderSection(employee_id, work_date, logs)
        main_layout.addWidget(header_widget)

        # Statistics
        stats_widget = self.createStatsSection(logs)
        main_layout.addWidget(stats_widget)

        # Table
        table_widget = self.createTableSection(logs)
        main_layout.addWidget(table_widget, 1)

        # Buttons
        button_widget = self.createButtonSection()
        main_layout.addWidget(button_widget)

        self.setLayout(main_layout)

    def createHeaderSection(self, employee_id, work_date, logs):
        header_widget = QWidget()
        header_layout = QVBoxLayout()
        header_layout.setSpacing(10)
        header_layout.setContentsMargins(0, 0, 0, 0)

        # Title
        title_label = QLabel("CHI TIẾT CHẤM CÔNG")
        title_label.setObjectName("titleLabel")
        title_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(title_label)

        # Info row
        info_layout = QHBoxLayout()
        info_layout.setSpacing(10)

        # Employee info
        employee_label = QLabel(f"👤 Nhân viên: {employee_id}")
        employee_label.setObjectName("infoLabel")

        # Date info
        date_label = QLabel(f"📅 Ngày: {work_date}")
        date_label.setObjectName("infoLabel")

        # Records count
        records_label = QLabel(f"📊 Số bản ghi: {len(logs)}")
        records_label.setObjectName("infoLabel")

        info_layout.addWidget(employee_label)
        info_layout.addWidget(date_label)
        info_layout.addWidget(records_label)
        info_layout.addStretch()

        header_layout.addLayout(info_layout)
        header_widget.setLayout(header_layout)

        return header_widget

    def createStatsSection(self, logs):
        stats_widget = QWidget()
        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(10)
        stats_layout.setContentsMargins(0, 0, 0, 0)

        # Calculate statistics
        stats = self.calculateStats(logs)

        for stat_name, stat_value, color in stats:
            card = QFrame()
            card.setObjectName("statsCard")
            card_layout = QVBoxLayout()
            card_layout.setSpacing(5)
            card_layout.setContentsMargins(10, 10, 10, 10)

            # Title
            title_label = QLabel(stat_name)
            title_label.setObjectName("statTitle")
            title_label.setAlignment(Qt.AlignCenter)

            # Value
            value_label = QLabel(str(stat_value))
            value_label.setObjectName("statValue")
            value_label.setAlignment(Qt.AlignCenter)
            if color:
                value_label.setStyleSheet(f"color: {color};")

            card_layout.addWidget(title_label)
            card_layout.addWidget(value_label)
            card.setLayout(card_layout)

            stats_layout.addWidget(card)

        stats_widget.setLayout(stats_layout)
        return stats_widget

    def calculateStats(self, logs):
        total_logs = len(logs)
        check_in = sum(1 for log in logs if log[5] and 'IN' in str(log[5]).upper())
        check_out = sum(1 for log in logs if log[5] and 'OUT' in str(log[5]).upper())

        # Calculate average confidence
        confidences = []
        for log in logs:
            try:
                if log[7] is not None:
                    confidence = float(log[7])
                    confidences.append(confidence)
            except (ValueError, TypeError):
                continue

        avg_confidence = sum(confidences) / len(confidences) if confidences else 0

        # Determine confidence color
        if avg_confidence >= 90:
            conf_color = "#28a745"  # Green
        elif avg_confidence >= 70:
            conf_color = "#ffc107"  # Yellow
        else:
            conf_color = "#dc3545"  # Red

        return [
            ("Tổng bản ghi", total_logs, "#007bff"),
            ("Check In", check_in, "#28a745"),
            ("Check Out", check_out, "#dc3545"),
            ("Độ tin cậy TB", f"{avg_confidence:.1f}%", conf_color)
        ]

    def createTableSection(self, logs):
        table_widget = QWidget()
        table_layout = QVBoxLayout()
        table_layout.setSpacing(10)
        table_layout.setContentsMargins(0, 0, 0, 0)

        # Table title
        table_title = QLabel("📋 Danh sách chi tiết")
        table_title.setStyleSheet("""
            font-size: 14px;
            font-weight: bold;
            color: #495057;
            padding: 5px 0px;
        """)
        table_layout.addWidget(table_title)

        # Create table
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "ID", "Thời gian", "Trạng thái", "Độ tin cậy", "Ảnh", "Thao tác"
        ])

        # Table setup
        table.setAlternatingRowColors(True)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setShowGrid(True)

        # Header configuration
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Fixed)
        header.setSectionResizeMode(3, QHeaderView.Fixed)
        header.setSectionResizeMode(4, QHeaderView.Fixed)
        header.setSectionResizeMode(5, QHeaderView.Fixed)

        # Column widths
        table.setColumnWidth(0, 60)
        table.setColumnWidth(2, 100)
        table.setColumnWidth(3, 100)
        table.setColumnWidth(4, 80)
        table.setColumnWidth(5, 90)

        # Populate table
        table.setRowCount(len(logs))
        table.verticalHeader().setDefaultSectionSize(40)
        table.verticalHeader().hide()

        for row, log in enumerate(logs):
            # ID
            id_item = QTableWidgetItem(str(log[0]))
            id_item.setTextAlignment(Qt.AlignCenter)
            id_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            table.setItem(row, 0, id_item)

            # Time
            time_item = QTableWidgetItem(str(log[3]) if log[3] else "N/A")
            time_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
            table.setItem(row, 1, time_item)

            # Status
            status_text = str(log[5]) if log[5] else "N/A"
            status_item = QTableWidgetItem(status_text)
            status_item.setTextAlignment(Qt.AlignCenter)
            status_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

            # Status coloring
            if 'IN' in status_text.upper():
                status_item.setBackground(QColor(40, 167, 69, 30))
                status_item.setForeground(QColor(40, 167, 69))
            elif 'OUT' in status_text.upper():
                status_item.setBackground(QColor(220, 53, 69, 30))
                status_item.setForeground(QColor(220, 53, 69))

            table.setItem(row, 2, status_item)

            # Confidence
            try:
                confidence = float(log[7]) if log[7] else 0
            except (ValueError, TypeError):
                confidence = 0

            confidence_item = QTableWidgetItem(f"{confidence:.1f}%")
            confidence_item.setTextAlignment(Qt.AlignCenter)
            confidence_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

            # Confidence coloring
            if confidence >= 90:
                confidence_item.setForeground(QColor(40, 167, 69))
            elif confidence >= 70:
                confidence_item.setForeground(QColor(255, 193, 7))
            else:
                confidence_item.setForeground(QColor(220, 53, 69))

            table.setItem(row, 3, confidence_item)

            # Image indicator
            image_path = log[6]
            has_image = bool(image_path) and os.path.exists(image_path)
            image_item = QTableWidgetItem("✓" if has_image else "✗")
            image_item.setTextAlignment(Qt.AlignCenter)
            image_item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

            if has_image:
                image_item.setForeground(QColor(40, 167, 69))
            else:
                image_item.setForeground(QColor(108, 117, 125))

            table.setItem(row, 4, image_item)

            # Action button
            view_btn = QPushButton("Xem")
            view_btn.setProperty("class", "view")
            view_btn.clicked.connect(lambda _, path=image_path: self.viewImage(path))

            if not has_image:
                view_btn.setEnabled(False)

            table.setCellWidget(row, 5, view_btn)

        table_layout.addWidget(table, 1)
        table_widget.setLayout(table_layout)

        return table_widget

    def createButtonSection(self):
        button_widget = QWidget()
        button_layout = QHBoxLayout()
        button_layout.setContentsMargins(0, 10, 0, 0)

        # Export button
        export_btn = QPushButton("📤 Xuất Excel")
        export_btn.setProperty("class", "export")
        export_btn.clicked.connect(self.exportToExcel)

        # Spacer
        button_layout.addStretch()

        # Close button
        close_btn = QPushButton("✖ Đóng")
        close_btn.setProperty("class", "close")
        close_btn.clicked.connect(self.close)

        button_layout.addWidget(export_btn)
        button_layout.addWidget(close_btn)

        button_widget.setLayout(button_layout)
        return button_widget

    def viewImage(self, image_path):
        """Hiển thị ảnh trong dialog đơn giản"""
        try:
            if image_path and os.path.exists(image_path):
                dialog = QDialog(self)
                dialog.setWindowTitle("Xem ảnh chấm công")
                dialog.resize(400, 300)
                dialog.setModal(True)

                # Simple styling
                dialog.setStyleSheet("""
                    QDialog {
                        background-color: white;
                    }
                    QLabel {
                        border: 1px solid #dee2e6;
                        border-radius: 8px;
                        padding: 10px;
                        background-color: #f8f9fa;
                    }
                    QPushButton {
                        background-color: #6c757d;
                        color: white;
                        border: none;
                        padding: 8px 16px;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #5a6268;
                    }
                """)

                layout = QVBoxLayout()
                layout.setSpacing(15)
                layout.setContentsMargins(20, 20, 20, 20)

                # Image
                label = QLabel()
                pixmap = QPixmap(image_path)
                if not pixmap.isNull():
                    scaled_pixmap = pixmap.scaled(350, 200, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    label.setPixmap(scaled_pixmap)
                    label.setAlignment(Qt.AlignCenter)
                else:
                    label.setText("❌ Không thể hiển thị ảnh")
                    label.setAlignment(Qt.AlignCenter)

                layout.addWidget(label)

                # Path
                path_label = QLabel(f"📁 {os.path.basename(image_path)}")
                path_label.setStyleSheet("color: #6c757d; font-size: 11px; border: none; padding: 5px;")
                layout.addWidget(path_label)

                # Close button
                close_btn = QPushButton("Đóng")
                close_btn.clicked.connect(dialog.close)
                layout.addWidget(close_btn, 0, Qt.AlignCenter)

                dialog.setLayout(layout)
                dialog.exec_()
            else:
                QMessageBox.information(self, "Thông báo", "Không có ảnh để hiển thị.")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi mở ảnh: {str(e)}")

    def exportToExcel(self):
        """Export to Excel with simple formatting"""
        if openpyxl is None:
            QMessageBox.warning(self, "Thiếu thư viện",
                                "Cần cài đặt thư viện openpyxl để xuất Excel.\n\n"
                                "Chạy lệnh: pip install openpyxl")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Lưu file Excel", "chi_tiet_cham_cong.xlsx", "Excel Files (*.xlsx)"
            )

            if file_path:
                wb = Workbook()
                ws = wb.active
                ws.title = "Chi tiết chấm công"

                # Headers
                headers = ["ID", "Mã NV", "Họ tên", "Thời gian", "Ngày", "Trạng thái", "Có ảnh", "Độ tin cậy"]

                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")

                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Write data
                for row, log in enumerate(self.logs, 2):
                    ws.cell(row=row, column=1, value=log[0])  # ID
                    ws.cell(row=row, column=2, value=log[1])  # Employee ID
                    ws.cell(row=row, column=3, value=log[2])  # Full Name
                    ws.cell(row=row, column=4, value=str(log[3]) if log[3] else "")  # Time
                    ws.cell(row=row, column=5, value=str(log[4]) if log[4] else "")  # Date
                    ws.cell(row=row, column=6, value=log[5])  # Status
                    ws.cell(row=row, column=7, value="Có" if (log[6] and os.path.exists(log[6])) else "Không")  # Image
                    ws.cell(row=row, column=8, value=f"{float(log[7]):.1f}%" if log[7] else "N/A")  # Confidence

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(file_path)
                QMessageBox.information(self, "Thành công", f"Đã xuất dữ liệu thành công!\n\nFile: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi xuất Excel: {str(e)}")
